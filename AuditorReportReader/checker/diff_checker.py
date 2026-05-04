"""
diff_checker.py — Compare a tool-produced filled Excel against a correct sample.

Outputs a structured diff with accuracy score and mismatch categories:
  MISSING      — filled has nothing, correct has a value
  SIGN_FLIP    — same magnitude but opposite sign
  SCALE_x10    — off by 10× (likely '000 misdetection)
  SCALE_x1000  — off by 1000× (RM'000 not applied)
  ROUNDING     — within 0.5% (acceptable)
  WRONG_VALUE  — genuinely different number
  TYPE_MISMATCH— one is text, other is numeric

Usage (standalone):
  python diff_checker.py <filled.xlsx> <correct.xlsx> [case_name]
"""

import re
import sys
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.utils import get_column_letter

_MAIN_SHEET = "Summary of Information"

# Known alternative sheet names (checked case-insensitively before scanning)
_KNOWN_SHEET_ALIASES = [
    "summary of information",
    "summary",
    "financial summary",
    "financial statements",
    "financial statement",
    "income statement",
    "profit & loss",
    "p&l",
    "pnl",
]

# Labels that strongly indicate a financial summary sheet (matched in col B)
_FINANCIAL_SIGNALS = [
    "revenue", "cost of sales", "gross profit", "net profit",
    "total asset", "total liab", "equity", "depreciation",
]


def detect_sheet(wb: openpyxl.Workbook) -> str:
    """
    Find the sheet in a workbook that contains financial statement data.

    Resolution order:
    1. Exact match on 'Summary of Information'
    2. Case-insensitive match on any known alias
    3. Sheet with the most financial label hits in column B
    4. First sheet as last resort

    Returns the sheet name (never raises).
    """
    sheets = wb.sheetnames

    # 1. Exact match
    if _MAIN_SHEET in sheets:
        return _MAIN_SHEET

    # 2. Case-insensitive alias match (preserves first match order)
    lower_to_actual = {s.lower(): s for s in sheets}
    for alias in _KNOWN_SHEET_ALIASES:
        if alias in lower_to_actual:
            return lower_to_actual[alias]

    # 3. Score every sheet by how many financial labels appear in column B
    best_name  = sheets[0]
    best_score = 0
    for name in sheets:
        ws    = wb[name]
        score = 0
        for row in ws.iter_rows(min_col=2, max_col=2, values_only=True):
            cell = row[0]
            if cell:
                lower = str(cell).lower()
                score += sum(1 for sig in _FINANCIAL_SIGNALS if sig in lower)
        if score > best_score:
            best_score = score
            best_name  = name

    return best_name

# All fields we check, in order: (col B label in template, short key)
_FIELDS: List[Tuple[str, str]] = [
    # ── Income Statement ──────────────────────────────────────────────────────
    ("Revenue",                           "revenue"),
    ("Cost of sales",                     "cos"),
    ("Gross Profit",                       "gross_profit"),
    ("Depreciation",                       "depreciation"),
    ("Staff Cost",                         "staff_cost"),
    ("Other Operating Expenses",           "other_op_exp"),
    ("Add : Other Income",                 "other_income"),
    ("Interest",                           "interest"),
    ("Taxes",                              "taxes"),
    ("Net Profit (Loss) for the Year",     "net_profit"),
    # ── Balance Sheet — Assets ────────────────────────────────────────────────
    ("Non Current Asset",                  "nca"),
    ("Current Asset",                      "ca"),
    ("Trade Receivables",                  "trade_rec"),
    ("Other Receivables and Prepayments",  "other_rec"),
    ("Amount Due from Directors",          "due_from_dir"),
    ("Amount Due from Related Companies",  "due_from_related"),
    ("Others",                             "others_ca"),
    ("Stock",                              "stock"),
    ("Cash & Cash At Bank",                "cash"),
    ("Total Asset",                        "total_asset"),
    # ── Balance Sheet — Liabilities ──────────────────────────────────────────
    ("Non Current Liabilities",            "ncl"),
    ("Current Liabilities",                "cl"),
    ("Trade Payables",                     "trade_pay"),
    ("Other Payables & Accruals",          "other_pay"),
    ("Amount Due to Director",             "due_to_dir"),
    ("Total Liabilities",                  "total_liab"),
    # ── Equity ───────────────────────────────────────────────────────────────
    ("Equity",                             "equity"),
    ("Share Capital",                      "share_cap"),
    ("Retained Earning",                   "retained"),
    ("Total Liabilities and Equity",       "total_l_e"),
]


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def _norm_label(text: str) -> str:
    return re.sub(r"\s+", " ", str(text).strip().lower())


def _to_float(v) -> Optional[float]:
    if v is None:
        return None
    s = str(v).replace(",", "").strip()
    if s.upper() in ("", "NOT FOUND", "NONE", "N/A", "TRUE", "FALSE"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _norm_val(v) -> str:
    if v is None:
        return ""
    f = _to_float(v)
    if f is not None:
        return str(int(round(f))) if abs(f - round(f)) < 0.005 else f"{f:.2f}"
    return str(v).strip()


def _classify(f_val, c_val) -> str:
    """Return error category for a mismatch."""
    c_norm = _norm_val(c_val)
    f_norm = _norm_val(f_val)

    if f_val in (None, "", "NOT FOUND") and c_norm not in ("", "0"):
        return "MISSING"

    f = _to_float(f_val)
    c = _to_float(c_val)

    if f is None or c is None:
        return "TYPE_MISMATCH"

    if c == 0:
        return "WRONG_VALUE"

    ratio = abs(f / c)

    # Sign flip: f ≈ –c
    if abs(f + c) < max(1.0, abs(c) * 0.005):
        return "SIGN_FLIP"

    # Scale off by 10×
    if 0.099 < ratio < 0.101 or 9.99 < ratio < 10.01:
        return "SCALE_x10"

    # Scale off by 100×
    if 0.0099 < ratio < 0.0101 or 99.9 < ratio < 100.1:
        return "SCALE_x100"

    # Scale off by 1000× (RM'000 not multiplied)
    if 0.00099 < ratio < 0.00101 or 999 < ratio < 1001:
        return "SCALE_x1000"

    # Rounding / close enough (within 0.5%)
    if 0.995 < ratio < 1.005:
        return "ROUNDING"

    return "WRONG_VALUE"


def _build_label_map(ws) -> Dict[str, List[int]]:
    m: Dict[str, List[int]] = {}
    for row in ws.iter_rows(min_col=2, max_col=2):
        cell = row[0]
        if cell.value and str(cell.value).strip():
            key = _norm_label(str(cell.value))
            m.setdefault(key, []).append(cell.row)
    return m


def _find_year_cols(ws) -> Dict[str, int]:
    """Return {year_str: col_index} by scanning row 4 for year patterns."""
    year_cols: Dict[str, int] = {}
    for col in range(3, ws.max_column + 1):
        raw = ws.cell(4, col).value
        if raw is None:
            continue
        val = str(raw)
        # Skip formula cells
        if val.startswith("="):
            continue
        m = re.search(r"(20\d{2})", val)
        if m:
            year = m.group(1)
            if year not in year_cols:   # first column wins for a given year
                year_cols[year] = col
    return year_cols


# ---------------------------------------------------------------------------
# main compare function
# ---------------------------------------------------------------------------

def compare(filled_path: str, correct_path: str,
            filled_sheet: Optional[str] = None,
            correct_sheet: Optional[str] = None) -> dict:
    """
    Compare filled Excel vs correct sample Excel.

    Sheet names are auto-detected if not supplied — 'Summary of Information'
    is tried first, then known aliases, then the sheet with the most
    financial labels in column B.

    Returns structured dict with per-field results and summary.
    """
    wb_f = openpyxl.load_workbook(filled_path, data_only=True)
    wb_c = openpyxl.load_workbook(correct_path, data_only=True)

    sheet_f = filled_sheet  or detect_sheet(wb_f)
    sheet_c = correct_sheet or detect_sheet(wb_c)

    if sheet_f not in wb_f.sheetnames:
        return {"error": f"Sheet '{sheet_f}' not found in {filled_path}. "
                         f"Available: {wb_f.sheetnames}"}
    if sheet_c not in wb_c.sheetnames:
        return {"error": f"Sheet '{sheet_c}' not found in {correct_path}. "
                         f"Available: {wb_c.sheetnames}"}

    if sheet_f != _MAIN_SHEET or sheet_c != _MAIN_SHEET:
        print(f"    [sheets]   filled='{sheet_f}'  correct='{sheet_c}'")

    ws_f = wb_f[sheet_f]
    ws_c = wb_c[sheet_c]

    years_f = _find_year_cols(ws_f)
    years_c = _find_year_cols(ws_c)

    # Only compare years where the filled file has actual numeric data
    # (skip template columns that exist as headers but were never populated)
    lm_f_pre = _build_label_map(ws_f)
    lm_c_pre = _build_label_map(ws_c)
    populated_years_f = set()
    for year, col in years_f.items():
        numeric_count = sum(
            1 for label, _ in _FIELDS
            if lm_f_pre.get(_norm_label(label))
            and _to_float(ws_f.cell(lm_f_pre[_norm_label(label)][0], col).value) is not None
        )
        if numeric_count >= 3:
            populated_years_f.add(year)

    common_years = sorted(set(populated_years_f) & set(years_c))

    lm_f = lm_f_pre
    lm_c = lm_c_pre

    fields_out = []
    total = matched = 0
    categories: Dict[str, List[str]] = {}

    for year in common_years:
        col_f = years_f[year]
        col_c = years_c[year]

        for label, field_key in _FIELDS:
            norm = _norm_label(label)
            rows_c = lm_c.get(norm, [])
            if not rows_c:
                continue

            c_val = ws_c.cell(rows_c[0], col_c).value
            # Skip if correct file has no expected value
            if c_val in (None, "", 0, 0.0):
                continue

            rows_f = lm_f.get(norm, [])
            f_val = ws_f.cell(rows_f[0], col_f).value if rows_f else None

            total += 1

            f_num = _to_float(f_val)
            c_num = _to_float(c_val)

            is_match = False
            if f_num is not None and c_num is not None:
                is_match = abs(f_num - c_num) < max(1.0, abs(c_num) * 0.005)
            elif _norm_val(f_val) == _norm_val(c_val):
                is_match = True

            if is_match:
                matched += 1
                status = "MATCH"
            else:
                status = _classify(f_val, c_val)

            categories.setdefault(status, []).append(f"{year}:{label}")
            fields_out.append({
                "year":    year,
                "label":   label,
                "field":   field_key,
                "filled":  f_val,
                "correct": c_val,
                "status":  status,
            })

    score_pct = round(100 * matched / total, 1) if total else 0.0

    return {
        "years_in_filled":  sorted(years_f),
        "years_in_correct": sorted(years_c),
        "years_compared":   common_years,
        "fields":           fields_out,
        "summary": {
            "total":           total,
            "matched":         matched,
            "score_pct":       score_pct,
            "by_category":     {k: len(v) for k, v in categories.items()},
            "fields_by_category": categories,
        },
    }


# ---------------------------------------------------------------------------
# report printer
# ---------------------------------------------------------------------------

def print_report(results: dict, case_name: str = ""):
    if "error" in results:
        print(f"  ERROR: {results['error']}")
        return

    s = results["summary"]
    bar = "=" * 64

    print(f"\n{bar}")
    if case_name:
        print(f"  CASE : {case_name}")
    print(f"  Years (filled)  : {results['years_in_filled']}")
    print(f"  Years (correct) : {results['years_in_correct']}")
    print(f"  Years compared  : {results['years_compared']}")
    score_color = "✓" if s["score_pct"] >= 80 else ("~" if s["score_pct"] >= 50 else "✗")
    print(f"  Score           : {s['matched']}/{s['total']}  ({s['score_pct']}%)  {score_color}")
    print(bar)

    # Print each error category
    ERROR_ORDER = ["MISSING", "SIGN_FLIP", "WRONG_VALUE",
                   "SCALE_x10", "SCALE_x100", "SCALE_x1000",
                   "TYPE_MISMATCH", "ROUNDING"]
    for cat in ERROR_ORDER:
        entries = [f for f in results["fields"] if f["status"] == cat]
        if not entries:
            continue
        print(f"\n  [{cat}]  ({len(entries)} fields)")
        for e in entries:
            fv = str(e["filled"]) if e["filled"] is not None else "—"
            cv = str(e["correct"])
            print(f"    {e['year']} | {e['label']:<42} "
                  f"filled={fv:<20} correct={cv}")

    if s["score_pct"] == 100:
        print("\n  ✓ Perfect match — no mismatches found.")
    print()


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python diff_checker.py <filled.xlsx> <correct.xlsx> [case_name]")
        print("       python diff_checker.py --detect <excel.xlsx>")
        sys.exit(1)

    # Quick sheet detection helper
    if sys.argv[1] == "--detect":
        if len(sys.argv) < 3:
            print("Usage: python diff_checker.py --detect <excel.xlsx>")
            sys.exit(1)
        wb   = openpyxl.load_workbook(sys.argv[2], data_only=True)
        name = detect_sheet(wb)
        print(f"Detected sheet: '{name}'")
        print(f"All sheets:     {wb.sheetnames}")
        sys.exit(0)

    if len(sys.argv) < 3:
        print("Usage: python diff_checker.py <filled.xlsx> <correct.xlsx> [case_name]")
        sys.exit(1)

    results = compare(sys.argv[1], sys.argv[2])
    print_report(results, sys.argv[3] if len(sys.argv) > 3 else "")

    s = results.get("summary", {})
    print(f"Score: {s.get('score_pct', 0)}%  ({s.get('matched', 0)}/{s.get('total', 0)})")
