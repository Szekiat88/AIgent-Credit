"""
Excel output writer — label-based, flexible for any template version.

Discovers rows at runtime by scanning column B (no hardcoded row numbers).
Discovers columns by scanning row 4 date headers.

Fills three areas of the workbook:
  1. "Summary of Information" — P&L rows 6-19, BS rows 23-53, summary rows 69-96
  2. "auditor"  — Phase 1-3 audit checks
  3. "CreditFlag" — final recommendation + arithmetic validation
"""

import re
from typing import Optional, Dict, List, Tuple

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_MAIN_SHEET = "Summary of Information"
_AUDITOR_SHEET = "auditor"
_CREDIT_SHEET = "CreditFlag"
_TOKEN_SHEET = "Token Usage"

# Gemini 2.5 Flash-Lite pricing (USD per 1M tokens, May 2025)
_COST_INPUT_PER_1M  = 0.075
_COST_OUTPUT_PER_1M = 0.30


# ---------------------------------------------------------------------------
# style helpers
# ---------------------------------------------------------------------------

def _red_bold() -> Font:
    return Font(bold=True, color="FF0000")

def _green_bold() -> Font:
    return Font(bold=True, color="00AA00")

def _header_font() -> Font:
    return Font(bold=True, color="FFFFFF")

def _header_fill() -> PatternFill:
    return PatternFill("solid", fgColor="1F4E79")

def _warn_fill() -> PatternFill:
    return PatternFill("solid", fgColor="FFC7CE")

def _ok_fill() -> PatternFill:
    return PatternFill("solid", fgColor="C6EFCE")

def _thin_border() -> Border:
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


# ---------------------------------------------------------------------------
# label / column discovery
# ---------------------------------------------------------------------------

def _norm_label(text: str) -> str:
    """Lowercase + collapse whitespace for label comparison."""
    return re.sub(r'\s+', ' ', str(text).strip().lower())


def _build_label_map(ws) -> Dict[str, List[int]]:
    """Scan column B → {normalised_label: [row_numbers]} (multiple rows per label possible)."""
    label_map: Dict[str, List[int]] = {}
    for row in ws.iter_rows(min_col=2, max_col=2):
        cell = row[0]
        if cell.value and str(cell.value).strip():
            key = _norm_label(str(cell.value))
            label_map.setdefault(key, []).append(cell.row)
    return label_map


def _section_bounds(label_map: Dict[str, List[int]],
                    start_label: str, end_label: str) -> Tuple[int, int]:
    """Return (start_row, end_row) bounding a section by anchor labels."""
    starts = label_map.get(_norm_label(start_label), [])
    ends = label_map.get(_norm_label(end_label), [])
    s = starts[0] if starts else 0
    e = ends[0] if ends else 9999
    return (s, e)


def _row_for(label_map: Dict[str, List[int]], label: str,
             lo: int = 0, hi: int = 9999) -> Optional[int]:
    """Find the row for `label` within (lo, hi].  Returns first match."""
    key = _norm_label(label)
    for r in label_map.get(key, []):
        if lo < r <= hi:
            return r
    return None


def _find_year_column(ws, target_year: str,
                      year_end_date: str = "",
                      preferred_col: int = 0) -> int:
    """
    Resolve which column to write for this year.

    Priority:
    1. Exact match on year_end_date in row 4 (e.g. '31/12/2022')
    2. Substring match on target_year (e.g. '2022') in row 4
    3. preferred_col if supplied — update its row-4 header with year_end_date
    4. First column (C onward) whose row-4 value is empty or is a formula ref
    5. Column C as hard fallback

    Always writes year_end_date (or target_year) to row 4 of the chosen column so
    the header reflects the actual report date.
    """
    label = year_end_date or target_year  # what we'll stamp into the header

    # ── 1 & 2: scan row 4 for an existing match ──────────────────────────────
    for col in range(3, ws.max_column + 1):
        raw = ws.cell(row=4, column=col).value
        val = str(raw or "")
        if year_end_date and year_end_date in val:
            return col  # exact date match — no header update needed
        if target_year and target_year in val and not str(raw).startswith("="):
            return col  # year substring match — no header update needed

    # ── 3: use the caller's preferred column ─────────────────────────────────
    if preferred_col:
        if label:
            ws.cell(row=4, column=preferred_col, value=label)
        return preferred_col

    # ── 4: find first column whose row-4 header is empty or a formula ─────────
    for col in range(3, min(ws.max_column + 2, 10)):
        raw = ws.cell(row=4, column=col).value
        if raw is None or str(raw).startswith("="):
            if label:
                ws.cell(row=4, column=col, value=label)
            return col

    # ── 5: hard fallback ──────────────────────────────────────────────────────
    if label:
        ws.cell(row=4, column=3, value=label)
    return 3


# ---------------------------------------------------------------------------
# cell writer
# ---------------------------------------------------------------------------

def _write_cell(ws, row: int, col: int, value,
                bold: bool = False, red: bool = False,
                fill: Optional[PatternFill] = None):
    if row is None:
        return
    cell = ws.cell(row=row, column=col, value=value)
    if bold or red:
        cell.font = Font(bold=bold, color="FF0000" if red else "000000")
    if fill:
        cell.fill = fill
    cell.alignment = Alignment(wrap_text=True, vertical="center")


# ---------------------------------------------------------------------------
# main financial data fill
# ---------------------------------------------------------------------------

def fill_financial_data(ws, financial_data: dict, target_year: str,
                        opinion: str = "",
                        year_end_date: str = "",
                        preferred_col: int = 0):
    """
    Write extracted financial values into 'Summary of Information'.
    Uses label-based row discovery — works even if rows shift in new templates.
    year_end_date (e.g. '31/12/2022') is written to the row-4 column header.
    preferred_col is used when the year doesn't match any existing header.
    """
    col = _find_year_column(ws, target_year, year_end_date=year_end_date,
                            preferred_col=preferred_col)
    display_date = year_end_date or target_year
    print(f"[Excel] Writing to column {get_column_letter(col)} (year: {display_date})")

    lm = _build_label_map(ws)

    def get(field: str) -> Optional[float]:
        return (financial_data.get(field) or {}).get("value")

    def write(label: str, value, lo: int = 0, hi: int = 9999,
              red_if_none: bool = True):
        row = _row_for(lm, label, lo, hi)
        if row is None:
            return
        if value is not None:
            _write_cell(ws, row, col, round(float(value), 2))
        elif red_if_none:
            cell = ws.cell(row=row, column=col, value="NOT FOUND")
            cell.font = _red_bold()

    # ── Section boundary discovery ─────────────────────────────────────────
    pl_lo, pl_hi = _section_bounds(lm, "profit & loss", "balance sheet")
    bs_lo, bs_hi = _section_bounds(lm, "balance sheet", "intangible asset")
    # NCA liabilities: between "non current liabilities" and "current liabilities"
    nca_lo, nca_hi = _section_bounds(lm, "non current liabilities", "current liabilities")
    # CL: between "current liabilities" and "total liabilities"
    cl_lo, cl_hi = _section_bounds(lm, "current liabilities", "total liabilities")

    # ── Profit & Loss ──────────────────────────────────────────────────────
    rev = get("Revenue")
    cos = get("Cost of Sales")
    gp  = get("Gross Profit") or ((rev - cos) if (rev and cos) else None)
    dep = get("Depreciation")
    sc  = get("Staff Cost")
    ooe = get("Other Operating Expenses")
    oi  = get("Other Income")
    int_exp = get("Interest / Finance Expenses")
    tax = get("Taxes")
    np_ = get("Net Profit (Loss) for the Year")

    # Derived fields
    op_profit = None
    ebit      = None
    pbt       = None
    if all(v is not None for v in [gp, dep, sc, ooe]):
        op_profit = gp - dep - sc - ooe
    if op_profit is not None and oi is not None:
        ebit = op_profit + oi
    if np_ is not None and tax is not None:
        pbt = np_ + tax

    write("Revenue",                          rev,     pl_lo, pl_hi)
    write("Cost of sales",                    cos,     pl_lo, pl_hi)
    write("Gross Profit",                     gp,      pl_lo, pl_hi)
    write("Depreciation",                     dep,     pl_lo, pl_hi)
    write("Staff Cost",                       sc,      pl_lo, pl_hi)
    write("Other Operating Expenses",         ooe,     pl_lo, pl_hi)
    write("Operating Profit",                 op_profit, pl_lo, pl_hi, red_if_none=False)
    write("Add : Other Income",               oi,      pl_lo, pl_hi)
    write("EBIT",                             ebit,    pl_lo, pl_hi, red_if_none=False)
    write("Interest",                         int_exp, pl_lo, pl_hi)
    write("Taxes",                            tax,     pl_lo, pl_hi)
    write("Net Profit (Loss) for the Year",   np_,     pl_lo, pl_hi)
    write("Dividend Payout",                  0,       pl_lo, pl_hi, red_if_none=False)
    write("PBT",                              pbt,     pl_lo, pl_hi, red_if_none=False)

    # ── Balance Sheet — opinion row ────────────────────────────────────────
    opinion_row = _row_for(lm, "unqualified reports", bs_lo, bs_hi)
    if opinion_row:
        val = "Yes" if opinion == "UNQUALIFIED" else (opinion or "NOT FOUND")
        cell = ws.cell(row=opinion_row, column=col, value=val)
        cell.font = Font(color="00AA00") if val == "Yes" else _red_bold()

    # ── Balance Sheet — Assets ─────────────────────────────────────────────
    ca_lo, ca_hi = _section_bounds(lm, "current asset", "total asset")

    write("Non Current Asset",                    get("Non Current Asset"),      bs_lo, bs_hi)
    write("Current Asset",                        get("Current Asset"),           bs_lo, bs_hi)
    write("Trade Receivables",                    get("Trade Receivables"),       ca_lo, ca_hi)
    write("Other Receivables and Prepayments",    get("Other Receivables and Prepayments"), ca_lo, ca_hi)
    write("Other Receivables",                    get("Other Receivables"),       ca_lo, ca_hi, red_if_none=False)
    write("Amount Due from Directors",            get("Amount Due from Directors"), ca_lo, ca_hi, red_if_none=False)
    write("Amount Due from Related Companies",    get("Amount Due from Related Companies"), ca_lo, ca_hi, red_if_none=False)
    write("Others",                               get("Others"),                  ca_lo, ca_hi, red_if_none=False)
    write("Stock",                                get("Stock"),                   ca_lo, ca_hi)
    write("Cash & Cash At Bank",                  get("Cash & Cash At Bank"),     ca_lo, ca_hi)
    write("Total Asset",                          get("Total Asset"),             bs_lo, bs_hi)

    # ── Balance Sheet — NCA Liabilities ────────────────────────────────────
    write("Non Current Liabilities",              get("Non Current Liabilities"), bs_lo, bs_hi)
    write("Bank/FI/Non-FI borrowings",            get("NCA Bank Borrowings"),     nca_lo, nca_hi, red_if_none=False)
    write("Hire Purchase Payable",                get("NCA Hire Purchase"),        nca_lo, nca_hi, red_if_none=False)
    write("Other Payables",                       get("NCA Other Payables"),       nca_lo, nca_hi, red_if_none=False)

    # ── Balance Sheet — Current Liabilities ────────────────────────────────
    write("Current Liabilities",                  get("Current Liabilities"),      bs_lo, bs_hi)
    write("Trade Payables",                       get("Trade Payables"),           cl_lo, cl_hi)
    write("Other Payables & Accruals",            get("Other Payables & Accruals"),cl_lo, cl_hi)
    write("Other Payables",                       get("CL Other Payables"),        cl_lo, cl_hi, red_if_none=False)
    write("Amount Due to Director",               get("Amount Due to Director"),   cl_lo, cl_hi, red_if_none=False)
    write("Amount Due to Related Companies",      get("Amount Due to Related Companies"), cl_lo, cl_hi, red_if_none=False)
    write("Others",                               get("CL Others"),                cl_lo, cl_hi, red_if_none=False)
    write("Bank/FI/Non-FI borrowings",            get("CL Bank Borrowings"),       cl_lo, cl_hi, red_if_none=False)
    write("Hire Purchase Payable",                get("CL Hire Purchase"),          cl_lo, cl_hi, red_if_none=False)

    # ── Balance Sheet — Liabilities + Equity ──────────────────────────────
    eq_lo = cl_hi
    write("Total Liabilities",                    get("Total Liabilities"),        bs_lo, bs_hi)
    write("Equity",                               get("Equity"),                   bs_lo, bs_hi)
    write("Share Capital",                        get("Share Capital"),             bs_lo, bs_hi)
    write("Retained Earnings",                    get("Retained Earnings"),         bs_lo, bs_hi)
    write("Revaluation Reserve",                  get("Revaluation Reserve"),       bs_lo, bs_hi, red_if_none=False)
    write("Total Liabilities and Equity",         get("Total Liabilities and Equity"), bs_lo, bs_hi)

    # ── Second BS summary section (rows ~69-83) ────────────────────────────
    bs2_lo, bs2_hi = _find_second_bs_section(lm, bs_lo)
    if bs2_lo:
        # Opinion
        op2_row = _row_for(lm, "unqualified reports", bs2_lo, bs2_hi)
        if op2_row:
            val = "Yes" if opinion == "UNQUALIFIED" else (opinion or "NOT FOUND")
            cell = ws.cell(row=op2_row, column=col, value=val)
            cell.font = Font(color="00AA00") if val == "Yes" else _red_bold()

        for label, field in [
            ("Non Current Asset", "Non Current Asset"),
            ("Current Asset", "Current Asset"),
            ("Total Asset", "Total Asset"),
            ("Non Current Liabilities", "Non Current Liabilities"),
            ("Current Liabilities", "Current Liabilities"),
            ("Total Liabilities", "Total Liabilities"),
            ("Share Capital", "Share Capital"),
            ("Retained Earning", "Retained Earnings"),     # template has 'Retained Earning' (no s)
            ("Retained Earnings", "Retained Earnings"),
            ("Reserve", "Revaluation Reserve"),
            ("Equity", "Equity"),
            ("Total Liabilities and Equity", "Total Liabilities and Equity"),
        ]:
            row2 = _row_for(lm, label, bs2_lo, bs2_hi)
            if row2:
                val = get(field)
                if val is not None:
                    _write_cell(ws, row2, col, round(val, 2))

    # ── Third section: compact P&L summary (rows ~89-96) ──────────────────
    pl2_lo, pl2_hi = _find_third_pl_section(lm, bs2_lo or bs_hi)
    if pl2_lo:
        for label, val in [
            ("Revenue", rev),
            ("Net Profit (Loss) Before Tax", pbt),
            ("Net Profit (Loss) After Tax", np_),
            ("Net Dividend", 0),
            ("Miniority Interest", 0),  # template has typo "Miniority"
        ]:
            row3 = _row_for(lm, label, pl2_lo, pl2_hi)
            if row3 and val is not None:
                _write_cell(ws, row3, col, round(float(val), 2) if isinstance(val, (int, float)) else val)


def _find_second_bs_section(label_map: Dict[str, List[int]],
                             first_bs_end: int) -> Tuple[Optional[int], int]:
    """Find the second Balance Sheet section that starts after first_bs_end."""
    key = _norm_label("balance sheet")
    rows = [r for r in label_map.get(key, []) if r > first_bs_end]
    if not rows:
        return None, 9999
    lo = rows[0]
    # End: look for "Audited / Mgmt Account" section below it
    end_key = _norm_label("audited / mgmt account")
    ends = [r for r in label_map.get(end_key, []) if r > lo]
    hi = ends[0] if ends else lo + 30
    return lo, hi


def _find_third_pl_section(label_map: Dict[str, List[int]],
                            above_row: int) -> Tuple[Optional[int], int]:
    """Find the compact P&L summary section (Revenue, PBT, Net Profit)."""
    key = _norm_label("profit & loss")
    rows = [r for r in label_map.get(key, []) if r > above_row]
    if not rows:
        return None, 9999
    lo = rows[0]
    return lo, lo + 20


# ---------------------------------------------------------------------------
# auditor sheet
# ---------------------------------------------------------------------------

def _ensure_sheet(wb: Workbook, name: str):
    if name not in wb.sheetnames:
        wb.create_sheet(name)
    return wb[name]


def fill_auditor_sheet(wb: Workbook, audit_checks: dict,
                       blacklist_firm: dict, blacklist_accountant: dict,
                       target_year: str):
    ws = _ensure_sheet(wb, _AUDITOR_SHEET)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 60

    opinion = audit_checks.get("opinion", "NOT FOUND")

    ws["A1"] = opinion
    ws["A1"].font = Font(bold=True, size=14,
                         color="FF0000" if opinion != "UNQUALIFIED" else "00AA00")

    def row(r, label, value, flag: bool = False):
        ws.cell(r, 1, label).font = Font(bold=True)
        cell = ws.cell(r, 2, str(value) if value is not None else "NOT FOUND")
        if flag:
            cell.font = _red_bold()
            cell.fill = _warn_fill()
        ws.cell(r, 1).border = _thin_border()
        ws.cell(r, 2).border = _thin_border()

    ws.cell(3, 1, "AUDIT OPINION").font = _header_font()
    ws.cell(3, 1).fill = _header_fill()
    ws.merge_cells("A3:B3")

    row(4, "Opinion Classification", opinion,
        flag=opinion not in ("UNQUALIFIED", "NOT FOUND"))
    row(5, "True and Fair View", "YES" if audit_checks.get("true_and_fair") else "NO",
        flag=not audit_checks.get("true_and_fair"))
    row(6, "Evidence", audit_checks.get("opinion_evidence", ""))

    ws.cell(8, 1, "AUDITOR INFORMATION").font = _header_font()
    ws.cell(8, 1).fill = _header_fill()
    ws.merge_cells("A8:B8")

    row(9,  "Auditor Firm Name",   audit_checks.get("firm_name"))
    row(10, "Signing Accountant",  audit_checks.get("accountant_name"))
    row(11, "MIA / AF Number",     audit_checks.get("mia_number"))
    row(12, "Signature Date",      audit_checks.get("signature_date"))

    ws.cell(14, 1, "BLACKLIST CHECK").font = _header_font()
    ws.cell(14, 1).fill = _header_fill()
    ws.merge_cells("A14:B14")

    firm_status = blacklist_firm.get("status", "INCONCLUSIVE")
    acct_status = blacklist_accountant.get("status", "INCONCLUSIVE")
    row(15, "Firm — Blacklist Status",     firm_status,  flag=firm_status == "FLAGGED")
    row(16, "Firm — Search Query",         blacklist_firm.get("query", ""))
    row(17, "Firm — Evidence",             "; ".join(blacklist_firm.get("evidence", [])))
    row(18, "Accountant — Blacklist Status", acct_status, flag=acct_status == "FLAGGED")
    row(19, "Accountant — Search Query",   blacklist_accountant.get("query", ""))
    row(20, "Accountant — Evidence",       "; ".join(blacklist_accountant.get("evidence", [])))

    ws.cell(22, 1, "SIGNATURE CONSISTENCY").font = _header_font()
    ws.cell(22, 1).fill = _header_fill()
    ws.merge_cells("A22:B22")

    sig_dates = audit_checks.get("signature_dates", {})
    r = 23
    for label, date in sig_dates.items():
        row(r, label, date)
        r += 1
    row(r, "Consistency Check", audit_checks.get("signature_consistency"),
        flag="INCONSISTENT" in str(audit_checks.get("signature_consistency", "")))
    r += 2

    ws.cell(r, 1, "STATUTORY DECLARATION").font = _header_font()
    ws.cell(r, 1).fill = _header_fill()
    ws.merge_cells(f"A{r}:B{r}")
    r += 1

    stat = audit_checks.get("statutory_declaration", {})
    row(r, "COP Name", stat.get("cop_name")); r += 1
    row(r, "COP Date", stat.get("cop_date")); r += 1
    row(r, "Date Consistent", "YES" if stat.get("date_consistent") else "NO",
        flag=not stat.get("date_consistent")); r += 1
    row(r, "Status", stat.get("status"), flag=stat.get("status") == "FLAGGED"); r += 1
    row(r, "Notes", stat.get("notes", ""))

    return opinion


# ---------------------------------------------------------------------------
# CreditFlag sheet
# ---------------------------------------------------------------------------

def fill_credit_flag_sheet(wb: Workbook, audit_checks: dict,
                            financial_data: dict,
                            blacklist_firm: dict,
                            blacklist_accountant: dict,
                            validation_results: Optional[dict] = None):
    ws = _ensure_sheet(wb, _CREDIT_SHEET)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 50

    def hdr(r, text):
        ws.cell(r, 1, text).font = _header_font()
        ws.cell(r, 1).fill = _header_fill()
        ws.merge_cells(f"A{r}:C{r}")

    def chk(r, label, status, notes=""):
        ws.cell(r, 1, label).font = Font(bold=True)
        cell = ws.cell(r, 2, status)
        ok = status in ("PASS", "CLEAN", "CONSISTENT", "VALID", "MATCH",
                        "UNQUALIFIED", "YES", "PROCEED")
        cell.font = _green_bold() if ok else _red_bold()
        cell.fill = _ok_fill() if ok else _warn_fill()
        ws.cell(r, 3, notes)
        for c in range(1, 4):
            ws.cell(r, c).border = _thin_border()

    hdr(1, "CREDIT ANALYST — FLAG SUMMARY")

    opinion     = audit_checks.get("opinion", "NOT FOUND")
    true_fair   = audit_checks.get("true_and_fair", False)
    firm_bl     = blacklist_firm.get("status", "INCONCLUSIVE")
    acct_bl     = blacklist_accountant.get("status", "INCONCLUSIVE")
    consistency = audit_checks.get("signature_consistency", "INSUFFICIENT DATA")
    stat_status = audit_checks.get("statutory_declaration", {}).get("status", "FLAGGED")

    net_profit = financial_data.get("Net Profit (Loss) for the Year", {})
    tci_match  = net_profit.get("validated_vs_tci")
    tci_status = "MATCH" if tci_match is True else "MISMATCH" if tci_match is False else "NOT CHECKED"

    chk(3,  "Audit Opinion", opinion if opinion == "UNQUALIFIED" else f"⚠ {opinion}",
        audit_checks.get("opinion_evidence", "")[:100])
    chk(4,  "True and Fair View", "YES" if true_fair else "NO — NOT STATED")
    chk(5,  "Auditor Firm Blacklist", firm_bl, "; ".join(blacklist_firm.get("evidence", []))[:80])
    chk(6,  "Accountant Blacklist", acct_bl, "; ".join(blacklist_accountant.get("evidence", []))[:80])
    chk(7,  "Signature Consistency",
        "CONSISTENT" if consistency.startswith("CONSISTENT") else consistency)
    chk(8,  "Statutory Declaration", stat_status,
        audit_checks.get("statutory_declaration", {}).get("notes", ""))
    chk(9,  "Net Profit = Total Comprehensive Income", tci_status)

    # Arithmetic validation
    if validation_results:
        hdr(11, "ARITHMETIC VALIDATION")
        r = 12
        for check_name, res in validation_results.items():
            status = res.get("status", "SKIP")
            note = ("missing data" if status == "SKIP" else ""
                    if status == "PASS" else
                    (f"computed={res.get('computed', 0):,.0f}  "
                     f"reported={res.get('reported', 0):,.0f}  "
                     f"diff={res.get('diff_pct', 0):.2f}%"))
            chk(r, check_name, status, note)
            r += 1
        rec_start = r + 2
    else:
        rec_start = 11

    # Recommendation
    escalate = [
        opinion == "QUALIFIED",
        firm_bl == "FLAGGED",
        acct_bl == "FLAGGED",
        "INCONSISTENT" in consistency,
        stat_status == "FLAGGED",
        tci_status == "MISMATCH",
    ]
    decline = [opinion in ("ADVERSE", "DISCLAIMED")]

    if any(decline):
        recommendation = "DECLINE"
    elif any(escalate):
        recommendation = "ESCALATE"
    else:
        recommendation = "PROCEED"

    hdr(rec_start, "OVERALL RECOMMENDATION")
    rec_row = rec_start + 1
    ws.cell(rec_row, 1, recommendation)
    ws.cell(rec_row, 1).font = Font(bold=True, size=16,
                                    color="FF0000" if recommendation != "PROCEED" else "00AA00")
    ws.cell(rec_row, 1).fill = _ok_fill() if recommendation == "PROCEED" else _warn_fill()
    ws.merge_cells(f"A{rec_row}:C{rec_row}")

    return recommendation


# ---------------------------------------------------------------------------
# Token Usage sheet
# ---------------------------------------------------------------------------

def fill_token_usage_sheet(wb: Workbook, token_usage: dict):
    """Write a Token Usage sheet with per-call stats and estimated cost."""
    ws = _ensure_sheet(wb, _TOKEN_SHEET)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16

    headers = ["Section", "Prompt Tokens", "Output Tokens", "Total Tokens", "Est. Cost (USD)"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = _header_font()
        cell.fill = _header_fill()
        cell.border = _thin_border()

    r = 2
    total_prompt = total_output = total_total = 0
    total_cost = 0.0

    for section, usage in token_usage.items():
        prompt_t = usage.get("prompt_tokens", 0) or 0
        output_t = usage.get("output_tokens", 0) or 0
        total_t  = usage.get("total_tokens", 0) or (prompt_t + output_t)
        cost = (prompt_t * _COST_INPUT_PER_1M + output_t * _COST_OUTPUT_PER_1M) / 1_000_000

        ws.cell(r, 1, section).border = _thin_border()
        ws.cell(r, 2, prompt_t).border = _thin_border()
        ws.cell(r, 3, output_t).border = _thin_border()
        ws.cell(r, 4, total_t).border = _thin_border()
        ws.cell(r, 5, round(cost, 6)).border = _thin_border()

        total_prompt += prompt_t
        total_output += output_t
        total_total  += total_t
        total_cost   += cost
        r += 1

    # Total row
    for c, val in enumerate(
        ["TOTAL", total_prompt, total_output, total_total, round(total_cost, 6)], 1
    ):
        cell = ws.cell(r, c, val)
        cell.font = Font(bold=True)
        cell.fill = _ok_fill()
        cell.border = _thin_border()


# ---------------------------------------------------------------------------
# main entry point
# ---------------------------------------------------------------------------

def write_output(
    template_path: str,
    output_path: str,
    target_year: str,
    audit_checks: dict,
    financial_data: dict,
    blacklist_firm: dict,
    blacklist_accountant: dict,
    validation_results: Optional[dict] = None,
    prior_financial_data: Optional[dict] = None,
    prior_year: str = "",
    prior_validation_results: Optional[dict] = None,
    token_usage: Optional[dict] = None,
    year_end_date: str = "",
    prior_year_end_date: str = "",
):
    wb = openpyxl.load_workbook(template_path)

    opinion = audit_checks.get("opinion", "NOT FOUND")

    if _MAIN_SHEET in wb.sheetnames:
        ws_main = wb[_MAIN_SHEET]
        # Prior year first (earlier column — C=3); current year after (later column — D=4).
        # preferred_col is only used when no header matches; it ensures the two years
        # land in separate columns even if neither matches an existing header.
        if prior_financial_data and prior_year:
            fill_financial_data(ws_main, prior_financial_data, prior_year, opinion,
                                year_end_date=prior_year_end_date, preferred_col=3)
        fill_financial_data(ws_main, financial_data, target_year, opinion,
                            year_end_date=year_end_date, preferred_col=4)

    fill_auditor_sheet(wb, audit_checks, blacklist_firm,
                       blacklist_accountant, target_year)
    recommendation = fill_credit_flag_sheet(
        wb, audit_checks, financial_data,
        blacklist_firm, blacklist_accountant,
        validation_results=validation_results,
    )

    if token_usage:
        fill_token_usage_sheet(wb, token_usage)

    wb.save(output_path)
    print(f"[Excel] Saved → {output_path}")
    print(f"[Excel] Recommendation: {recommendation}")
    return recommendation
