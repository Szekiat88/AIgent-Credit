"""
Keyword intelligence layer.

Each financial field has:
  - primary   : keywords that unambiguously identify it
  - fallback  : broader terms used by some auditors
  - note_hint : additional terms that appear inside a referenced note

Users can extend this by adding rows to the "KeywordMap" sheet in the Excel
template:
    Column A = field name (must match a key in BUILT_IN_FIELDS)
    Column B = extra primary keyword(s), comma-separated
    Column C = extra fallback keyword(s), comma-separated
"""

import difflib
import re
from typing import Optional

# ---------------------------------------------------------------------------
# Built-in field definitions
# ---------------------------------------------------------------------------
# excel_row : 1-based row in "Summary of Information" sheet
# excel_col_label : human label (used for logging / auditor sheet)
# note_header : regex variants that head the notes block for this item

BUILT_IN_FIELDS: dict[str, dict] = {

    # ── Income Statement ────────────────────────────────────────────────────
    "Revenue": {
        "excel_row": 5,
        "primary": ["revenue", "turnover", "net revenue", "total revenue",
                    "sales", "net sales", "total sales"],
        "fallback": ["income from operations", "gross receipts"],
    },
    "Cost of Sales": {
        "excel_row": 6,
        "primary": ["cost of sales", "cost of goods sold", "cost of revenue",
                    "cost of services rendered", "cost of services"],
        "fallback": ["direct costs", "production costs", "cogs", "cost of works"],
    },
    "Gross Profit": {
        "excel_row": 7,
        "primary": ["gross profit", "gross income", "gross loss"],
        "fallback": [],
    },
    "Depreciation": {
        "excel_row": 8,
        "primary": ["depreciation of property", "depreciation of ppe",
                    "depreciation and amortisation",
                    "depreciation and amortization",
                    "depreciation of right-of-use",
                    "amortisation of right-of-use",
                    "depreciation"],
        "fallback": ["amortisation", "amortization", "write-off of ppe"],
        "note_header": r"depreciation|amortis",
    },
    "Staff Cost": {
        "excel_row": 9,
        "primary": ["staff costs", "staff cost", "employee benefits expense",
                    "employee costs", "personnel costs",
                    "remuneration and benefits"],
        "fallback": ["human resource", "hr costs", "manpower"],
        "note_header": r"staff cost|employee benefit",
        "sub_items": [
            # (sub-item label for logging, keywords)
            ("Directors Remuneration - Emoluments",
             ["directors' emoluments", "directors emoluments", "emoluments"]),
            ("Directors Remuneration - Fees",
             ["directors' fees", "directors fees", "fees"]),
            ("Directors EPF",
             ["directors' epf", "directors epf", "epf contribution - director"]),
            ("Directors SOCSO",
             ["directors' socso", "directors socso", "socso - director"]),
            ("Directors EIS",
             ["directors' eis", "directors eis", "eis - director"]),
            ("Wages Salaries Bonus Allowances",
             ["wages", "salaries", "bonus", "allowances",
              "wages, salaries", "wages and salaries",
              "salaries and wages", "salary"]),
            ("Sales Commission",
             ["sales commission", "commission"]),
            ("Staff Welfare",
             ["staff welfare", "employee welfare", "welfare"]),
            ("Staff Refreshment",
             ["refreshment", "staff refreshment"]),
            ("SOCSO",
             ["social security", "socso", "perkeso",
              "employer socso", "socso contribution"]),
            ("EIS",
             ["employment insurance", "eis", "sistem insurans pekerjaan",
              "eis contribution"]),
            ("EPF",
             ["employees provident fund", "epf", "kwsp",
              "provident fund", "epf contribution"]),
        ],
    },
    "Other Operating Expenses": {
        "excel_row": 10,
        "primary": ["other operating expenses", "administrative and other operating",
                    "administrative expenses", "general and administrative",
                    "selling and distribution", "operating expenses"],
        "fallback": ["overheads", "general expenses"],
        # derived: raw_admin - staff_cost - finance_expenses
        "derived": True,
    },
    "Other Income": {
        "excel_row": 12,
        "primary": ["other income", "other operating income",
                    "other revenue", "other gains"],
        "fallback": ["miscellaneous income", "sundry income"],
    },
    "Interest / Finance Expenses": {
        "excel_row": 14,
        "primary": ["finance costs", "finance expenses", "financial expenses",
                    "interest expenses", "borrowing costs",
                    "interest expense", "finance charges"],
        "fallback": ["bank charges", "interest on borrowings"],
        "note_header": r"finance cost|interest expense|borrowing cost",
        "sub_items": [
            ("Bank Overdraft Interest",
             ["bank overdraft interest", "overdraft interest",
              "interest on bank overdraft"]),
            ("Bank Acceptance Interest",
             ["bank acceptance", "ba interest", "trust receipt interest",
              "bankers acceptance"]),
            ("Hire Purchase Interest",
             ["hire purchase interest", "hp interest", "lease interest",
              "interest on hire purchase", "interest on lease"]),
            ("Term Loan Interest",
             ["term loan interest", "interest on term loan",
              "term loan", "interest - term"]),
            ("Revolving Credit Interest",
             ["revolving credit interest", "revolving credit",
              "interest on revolving", "rc interest"]),
        ],
        "catch_all_interest": True,   # also sum any line with "interest"
    },
    "Taxes": {
        "excel_row": 15,
        "primary": ["income tax expense", "tax expense", "taxation",
                    "income tax", "current tax", "deferred tax expense"],
        "fallback": ["zakat", "tax charge"],
    },
    "Net Profit (Loss) for the Year": {
        "excel_row": 16,
        "primary": ["profit/(loss) for the year",
                    "profit for the year",
                    "loss for the year",
                    "net profit for the year",
                    "net loss for the year",
                    "profit after tax",
                    "net profit after tax",
                    "net loss after tax",
                    "total comprehensive income for the year",
                    "total comprehensive loss for the year"],
        "fallback": ["net income", "net earnings", "bottom line"],
        # validation: must match "total comprehensive income"
        "validate_match": "total comprehensive income",
    },

    # ── Balance Sheet — Assets ───────────────────────────────────────────────
    "Non Current Asset": {
        "excel_row": 23,
        "primary": ["non-current assets", "non current assets",
                    "fixed assets", "long-term assets"],
        "fallback": ["property plant equipment total", "total ppe"],
    },
    "Current Asset": {
        "excel_row": 24,
        "primary": ["current assets", "total current assets"],
        "fallback": [],
    },
    "Trade Receivables": {
        "excel_row": 25,
        "primary": ["trade receivables", "trade debtors",
                    "trade and other receivables", "accounts receivable"],
        "fallback": ["debtors"],
    },
    "Other Receivables and Prepayments": {
        "excel_row": 26,
        "primary": ["other receivables and prepayments",
                    "other receivables, deposits and prepayments",
                    "deposits, prepayments and other receivables"],
        "fallback": [],
    },
    "Other Receivables": {
        "excel_row": 27,
        "primary": ["other receivables", "sundry debtors",
                    "sundry receivables"],
        "fallback": [],
        # derived from note: total_other_receivable - related_co - deposits - prepayments
        "derived": True,
    },
    "Amount Due from Directors": {
        "excel_row": 28,
        "primary": ["amount due from director", "due from director",
                    "amount owing by director", "director current account"],
        "fallback": ["director loan receivable"],
    },
    "Amount Due from Related Companies": {
        "excel_row": 29,
        "primary": ["amount due from related", "due from related compan",
                    "amount owing by related", "related party receivable",
                    "amount due from a related", "due from related parties"],
        "fallback": ["intercompany receivable", "intra-group receivable"],
    },
    "Others": {
        "excel_row": 30,
        "primary": ["deposits and prepayments", "deposits, prepayments",
                    "prepayments and deposits", "deposit and prepayment"],
        "fallback": ["deposits", "prepayments", "advances"],
    },
    "Stock": {
        "excel_row": 31,
        "primary": ["inventories", "stocks", "inventory",
                    "finished goods", "raw materials", "work in progress"],
        "fallback": ["merchandise", "goods on hand"],
    },
    "Cash & Cash At Bank": {
        "excel_row": 32,
        "primary": ["cash and bank balances", "cash and cash equivalents",
                    "cash at bank", "bank balances", "cash and bank",
                    "cash on hand and at bank"],
        "fallback": ["bank and cash", "fixed deposits"],
    },
    "Total Asset": {
        "excel_row": 33,
        "primary": ["total assets"],
        "fallback": [],
    },

    # ── Balance Sheet — Liabilities ──────────────────────────────────────────
    "Non Current Liabilities": {
        "excel_row": 34,
        "primary": ["non-current liabilities", "non current liabilities",
                    "long-term liabilities"],
        "fallback": [],
    },
    "NCA Bank Borrowings": {
        "excel_row": 35,
        "primary": ["term loan", "bank borrowings", "finance lease",
                    "long-term borrowings", "long term loan",
                    "bank and other borrowings"],
        "fallback": ["bank facilities", "borrowings"],
        "section": "non_current",
    },
    "NCA Hire Purchase": {
        "excel_row": 36,
        "primary": ["hire purchase payable", "lease liabilities",
                    "right-of-use liabilities", "finance lease payable"],
        "fallback": [],
        "section": "non_current",
    },
    "NCA Other Payables": {
        "excel_row": 37,
        "primary": ["other payables", "other creditors", "deferred income"],
        "fallback": [],
        "section": "non_current",
    },
    "Current Liabilities": {
        "excel_row": 38,
        "primary": ["current liabilities", "total current liabilities"],
        "fallback": [],
    },
    "Trade Payables": {
        "excel_row": 39,
        "primary": ["trade payables", "trade creditors",
                    "trade and other payables", "accounts payable"],
        "fallback": ["creditors"],
    },
    "Other Payables & Accruals": {
        "excel_row": 40,
        "primary": ["other payables and accruals", "other payables & accruals",
                    "accrued liabilities", "accruals and other payables"],
        "fallback": [],
    },
    "CL Other Payables": {
        "excel_row": 41,
        "primary": ["other payables", "sundry creditors", "sundry payables"],
        "fallback": [],
        "section": "current",
    },
    "Amount Due to Director": {
        "excel_row": 42,
        "primary": ["amount due to director", "due to director",
                    "amount owing to director", "director loan payable"],
        "fallback": [],
    },
    "Amount Due to Related Companies": {
        "excel_row": 43,
        "primary": ["amount due to related", "due to related compan",
                    "amount owing to related", "related party payable",
                    "due to related parties"],
        "fallback": ["intercompany payable"],
    },
    "CL Others": {
        "excel_row": 44,
        "primary": ["other payables", "deposits received", "advance from customer"],
        "fallback": [],
        "section": "current",
    },
    "CL Bank Borrowings": {
        "excel_row": 45,
        "primary": ["bank overdraft", "short-term borrowings", "bank borrowings",
                    "revolving credit", "banker acceptance",
                    "trust receipts", "short term loan"],
        "fallback": [],
        "section": "current",
    },
    "CL Hire Purchase": {
        "excel_row": 46,
        "primary": ["hire purchase payable", "lease liabilities",
                    "finance lease payable"],
        "fallback": [],
        "section": "current",
    },
    "Total Liabilities": {
        "excel_row": 47,
        "primary": ["total liabilities"],
        "fallback": [],
    },

    # ── Equity ───────────────────────────────────────────────────────────────
    "Equity": {
        "excel_row": 48,
        "primary": ["total equity", "shareholders equity",
                    "shareholders' equity", "owner's equity",
                    "total equity attributable"],
        "fallback": [],
    },
    "Share Capital": {
        "excel_row": 49,
        "primary": ["share capital", "paid-up capital", "issued capital",
                    "ordinary shares"],
        "fallback": ["capital"],
    },
    "Retained Earnings": {
        "excel_row": 50,
        "primary": ["retained earnings", "retained profits",
                    "accumulated profits", "accumulated losses",
                    "retained earnings/(accumulated losses)"],
        "fallback": ["revenue reserve"],
    },
    "Revaluation Reserve": {
        "excel_row": 51,
        "primary": ["revaluation reserve", "revaluation surplus",
                    "asset revaluation"],
        "fallback": ["other reserves"],
    },
    "Total Liabilities and Equity": {
        "excel_row": 52,
        "primary": ["total liabilities and equity",
                    "total equity and liabilities",
                    "total liabilities and shareholders",
                    "total equity and total liabilities"],
        "fallback": [],
    },
}

# Section header variants for the six key report sections
SECTION_HEADERS: dict[str, list[str]] = {
    "auditors_report": [
        "independent auditors' report",
        "independent auditor's report",
        "auditors' report",
        "report of the auditors",
        "auditor's report",
    ],
    "directors_report": [
        "directors' report",
        "director's report",
        "report of the directors",
        "directors report",
    ],
    "statement_by_directors": [
        "statement by directors",
        "directors' declaration",
        "statement of directors",
        "declaration by directors",
    ],
    "statutory_declaration": [
        "statutory declaration",
        "pengakuan berkanun",
    ],
    "income_statement": [
        "statement of comprehensive income",
        "statement of profit or loss",
        "statement of profit and loss",
        "income statement",
        "profit and loss account",
        "profit or loss",
    ],
    "balance_sheet": [
        "statement of financial position",
        "balance sheet",
        "penyata kedudukan kewangan",
    ],
    "notes": [
        "notes to the financial statements",
        "notes to financial statements",
        "notes to the accounts",
        "nota kepada penyata kewangan",
    ],
}

# ---------------------------------------------------------------------------
# fuzzy matching
# ---------------------------------------------------------------------------

def _norm(text: str) -> str:
    """Lowercase, collapse whitespace, strip punctuation for comparison."""
    text = text.lower()
    text = re.sub(r"['‘’“”]", "", text)
    text = re.sub(r"[/\\()\[\]]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def fuzzy_score(a: str, b: str) -> float:
    """0-1 similarity between two normalised strings."""
    return difflib.SequenceMatcher(None, _norm(a), _norm(b)).ratio()


def best_keyword_match(line: str, field_name: str,
                        user_map: Optional[dict] = None) -> tuple[float, str]:
    """
    Return (confidence 0-100, matched_keyword) for the best keyword hit
    in `line` for the given field_name.

    Strategy:
      1. Exact substring match in primary   → 95
      2. Exact substring match in fallback  → 75
      3. User-defined primary               → 90
      4. User-defined fallback              → 70
      5. Fuzzy match (>= 0.80)             → scaled 60-85
    """
    field = BUILT_IN_FIELDS.get(field_name, {})
    norm_line = _norm(line)

    def _exact_in(keywords: list[str]) -> Optional[str]:
        for kw in keywords:
            if _norm(kw) in norm_line:
                return kw
        return None

    hit = _exact_in(field.get("primary", []))
    if hit:
        return 95.0, hit

    if user_map:
        user_primary = user_map.get(field_name, {}).get("primary", [])
        hit = _exact_in(user_primary)
        if hit:
            return 90.0, hit

    hit = _exact_in(field.get("fallback", []))
    if hit:
        return 75.0, hit

    if user_map:
        user_fallback = user_map.get(field_name, {}).get("fallback", [])
        hit = _exact_in(user_fallback)
        if hit:
            return 70.0, hit

    # fuzzy fallback: check every primary keyword
    best_score, best_kw = 0.0, ""
    for kw in field.get("primary", []) + field.get("fallback", []):
        s = fuzzy_score(kw, norm_line[:len(kw) + 20])  # limit comparison window
        if s > best_score:
            best_score, best_kw = s, kw

    if best_score >= 0.80:
        return round(best_score * 85, 1), best_kw

    return 0.0, ""


# ---------------------------------------------------------------------------
# load user-defined keyword map from Excel
# ---------------------------------------------------------------------------

def load_user_keyword_map(excel_path: str) -> dict:
    """
    Read the 'KeywordMap' sheet (if present) from the Excel workbook.
    Returns dict: {field_name: {"primary": [...], "fallback": [...]}}
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        if "KeywordMap" not in wb.sheetnames:
            return {}
        ws = wb["KeywordMap"]
        user_map: dict = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            field = str(row[0]).strip()
            primary = [k.strip() for k in str(row[1] or "").split(",") if k.strip()]
            fallback = [k.strip() for k in str(row[2] or "").split(",") if k.strip()]
            user_map[field] = {"primary": primary, "fallback": fallback}
        return user_map
    except Exception:
        return {}
