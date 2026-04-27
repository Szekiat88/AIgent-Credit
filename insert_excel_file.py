"""Fill Knockout Matrix Excel from merged credit report data."""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Sequence

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from merged_credit_report import merge_reports, resolve_pdf_path

from column_l_validator import apply_column_l_highlighting, RED_BOLD_FONT
from text_normalize import normalize_compare_text
from credit_analyst import assess, Assessment

SHEET_NAME = "Knock-Out"
LABEL_COL = 4  # Column D
DEFAULT_EXCEL = "Knockout Matrix Template.xlsx"
SCORE_RANGE_EQUIVALENTS = [
    (742, float("inf"), "A"), (701, 740, "A"), (661, 700, "B"),
    (621, 660, "B"), (581, 620, "C"), (541, 580, "C"),
    (501, 540, "D"), (461, 500, "E"), (421, 460, "F"), (0, 420, "F"),
]

# Row labels in column D (subject columns are issuer_data_col + 0, +2, +4, …).
LBL_SCORE_RAW = "Scoring by CRA Agency (Issuer's Credit Agency Score)"
LBL_SCORE_EQ = "Scoring by CRA Agency (Credit Score Equivalent)"
LBL_OPS_YEARS = (
    "Business has been in operations for at least THREE (3) years "
    "(Including upgrade from Sole Proprietorship and Partnership under similar business activity)"
)
LBL_COMPANY_STATUS = "Company Status (Existing Only)"
LBL_EXEMPT_PRIVATE = "Exempt Private Company"
LBL_WINDING_UP = "Winding Up / Bankruptcy Proceedings Record"
LBL_CREDIT_APPR = "Credit Applications Approved for Last 12 months (per primary CRA report)"
LBL_CREDIT_PEND = "Credit Applications Pending (per primary CRA report)"
LBL_LEGAL_ACTION = "Legal Action taken (from Banking) (per primary CRA report)"
LBL_EXISTING_FAC = "Existing No. of Facility (from Banking) (per primary CRA report)"
LBL_LEGAL_SUITS = "Legal Suits (per primary CRA report) (either as Plaintiff or Defendant)"
LBL_TRADE_CREDIT = "Trade / Credit Reference (per primary CRA report)"
LBL_LEGAL_CASE_STATUS = "Legal Case - Status (per primary CRA report)"
LBL_TOTAL_ENQ = (
    "Total Enquiries for Last 12 months (per primary CRA report) (Financial Related Search Count)"
)
LBL_SPECIAL_ATTN = "Special Attention Account (per primary CRA report)"
LBL_OVERDRAFT = (
    "Overdraft facility outstanding amount does not exceed the approved overdraft limit "
    "as per CCRIS (based on the primary CRA report)"
)
LBL_BANKING_WITHIN = (
    "Issuer's Total Banking Outstanding Facilities does not exceed the Total Banking Limit "
    "(per primary CRA report)"
)
LBL_SUM_OUT = "Summary of Total Liabilities (Outstanding) (per primary CRA report)"
LBL_SUM_LIM = "Summary of Total Liabilities (Total Limit) (per primary CRA report)"
LBL_NONBANK_WITHIN = (
    "Issuer's Total Non- Bank Lender Outstanding Facilities does not exceed the "
    "Total Non-Bank Lender Limit (per primary CRA report)"
)
LBL_CCRIS_CONDUCT = "CCRIS Loan Account - Conduct Count (per primary CRA report)"
LBL_CCRIS_LEGAL = "CCRIS Loan Account - Legal Status (per primary CRA report)"
LBL_NLCI_CONDUCT = "Non-Bank Lender Credit Information (NLCI)- Conduct Count (per primary CRA report)"
LBL_NLCI_LEGAL = "Non-Bank Lender Credit Information (NLCI) - Legal Status (per primary CRA report)"
LBL_TOTAL_LIMIT = "Total Limit"
LBL_TOTAL_OUTSTANDING = "Total Outstanding Balance"


@dataclass(frozen=True)
class KnockoutCellPlacement:
    """One Excel cell: row from label (column D), column = issuer_data_col + col_offset."""

    label: str
    col_offset: int
    value: Any


def _subject_col_offset(subject_index: int) -> int:
    """1-based subject index → horizontal offset (0, 2, 4, …)."""
    return (subject_index - 1) * 2


def _norm(s: str) -> str:
    """Normalize Knock-Out template label text for comparison."""
    return normalize_compare_text(s, smart_typography=True)


def _safe_int(value: Any) -> int:
    """Safely convert value to int."""
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def _format_ol_status(status: str, outstanding: Optional[float], limit: Optional[float]) -> str:
    return (
        f"{status}, outstanding: {_format_with_commas(outstanding)}, "
        f"limit: {_format_with_commas(limit)}"
    )


def _format_with_commas(value: Optional[int | float]) -> str:
    """Format numeric value with comma separators for thousands."""
    if value is None:
        return "N/A"
    if float(value).is_integer():
        return f"{int(value):,}"
    return f"{value:,.2f}".rstrip("0").rstrip(".")


def _format_number(value: Optional[float | int]) -> Optional[str]:
    """Format number for display."""
    if value is None:
        return None
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, (int, float)):
        return _format_with_commas(value)
    if isinstance(value, str):
        cleaned = value.strip().replace(",", "")
        if re.fullmatch(r"-?\d+(\.\d+)?", cleaned):
            return _format_with_commas(float(cleaned))
    return str(value)


def _format_number_or_na(value: Optional[float | int | str]) -> str:
    """Format number-like values and coerce missing/blank values to N/A."""
    formatted = _format_number(value)
    if formatted is None:
        return "N/A"
    if isinstance(formatted, str) and not formatted.strip():
        return "N/A"
    return formatted


def _get_freq(d: Dict[str, Any], key: str) -> Optional[Dict[str, Any]]:
    v = d.get(key)
    return v.get("freq") if isinstance(v, dict) else None


def _safe_get(d: Dict[str, Any], key: str, default: Any) -> Any:
    v = d.get(key)
    return v if isinstance(v, type(default)) else default


def _format_mia_bucket_line(counts_dict: Dict[str, Any], plus_key: str) -> str:
    """Format one MIA1..MIA4+ line (CCRIS digit buckets use '5_plus'; NLCI freq uses '4+')."""
    return (
        f"MIA1: {_format_with_commas(_safe_int(counts_dict.get('1', 0)))}, "
        f"MIA2: {_format_with_commas(_safe_int(counts_dict.get('2', 0)))}, "
        f"MIA3: {_format_with_commas(_safe_int(counts_dict.get('3', 0)))}, "
        f"MIA4+: {_format_with_commas(_safe_int(counts_dict.get(plus_key, 0)))}"
    )


def _format_mia_counts(value: Dict[str, Any]) -> Optional[str]:
    """Format MIA counts for display."""
    counts = {
        "next_six": value.get("next_six_numbers_digit_counts_0_1_2_3_5_plus"),
        "next_first": value.get("next_first_numbers_digit_counts_0_1_2_3_5_plus"),
        "last_1": _get_freq(value, "last_1_month"),
        "last_6": _get_freq(value, "last_6_months"),
    }
    
    if not any(isinstance(v, dict) for v in counts.values()):
        return None

    parts = []
    if isinstance(counts["next_six"], dict):
        parts.append(f"past 6 months {_format_mia_bucket_line(counts['next_six'], '5_plus')}")
    if isinstance(counts["last_6"], dict):
        parts.append(f"past 6 months {_format_mia_bucket_line(counts['last_6'], '4+')}")
    if isinstance(counts["next_first"], dict):
        parts.append(f"current 1 month {_format_mia_bucket_line(counts['next_first'], '5_plus')}")
    if isinstance(counts["last_1"], dict):
        parts.append(f"current 1 month {_format_mia_bucket_line(counts['last_1'], '4+')}")
   

    return " and /or ".join(parts) if parts else None


def _format_cell_value(value: Any) -> Any:
    """Format cell value for Excel insertion."""
    if isinstance(value, dict):
        print(f"⚠️ Warning: Complex dict value for cell, attempting to format: {value}")
        mia_counts = _format_mia_counts(value)
        print(f"⚠️ Formatted MIA counts: {mia_counts}")
        return mia_counts if mia_counts else json.dumps(value, ensure_ascii=False)
    if isinstance(value, list):
        return json.dumps(value, ensure_ascii=False)
    if isinstance(value, (int, float, str)):
        return _format_number(value)
    return value


def _compute_overdraft_compliance(analysis: Dict[str, Any]) -> str:
    """Compute overdraft compliance status."""
    overdraft_comparisons = analysis.get("overdraft_comparisons", {})

    if not overdraft_comparisons:
        return "N/A"

    total_outstanding = 0.0
    total_limit = 0.0
    all_within_limit = True

    for comparison in overdraft_comparisons.values():
        outstanding = comparison.get("outstanding")
        limit = comparison.get("limit")

        if outstanding is None or limit is None:
            all_within_limit = False
            continue

        total_outstanding += float(outstanding)
        total_limit += float(limit)
        if float(outstanding) > float(limit):
            all_within_limit = False

    status = "YES" if all_within_limit else "NO"
    return _format_ol_status(status, total_outstanding, total_limit)


def _merge_overdraft_comparisons(sections: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Merge overdraft comparisons across all detailed report sections."""
    merged: Dict[str, Dict[str, Optional[float]]] = {}

    for section in sections:
        analysis = section.get("account_line_analysis", {})
        comparisons = analysis.get("overdraft_comparisons", {})
        if not isinstance(comparisons, dict):
            continue

        section_number = section.get("section_number")
        for record_no, values in comparisons.items():
            if not isinstance(values, dict):
                continue

            key = f"{section_number}:{record_no}" if section_number is not None else str(record_no)
            merged[key] = {
                "outstanding": values.get("outstanding"),
                "limit": values.get("limit"),
            }

    return {"overdraft_comparisons": merged}


def _compute_banking_facility_status(analysis: Dict[str, Any]) -> tuple[str, Optional[float], Optional[float]]:
    """Compute per-record banking outstanding-vs-limit status and first valid values."""
    comparisons = analysis.get("outstanding_limit_comparisons", {})
    if not comparisons:
        return "N/A", None, None

    entries: List[str] = []
    section_outstanding: Optional[float] = None
    section_limit: Optional[float] = None
    for record_key in sorted(comparisons.keys(), key=str):
        values = comparisons.get(record_key)
        if not isinstance(values, dict):
            continue
        outstanding = values.get("outstanding")
        limit = values.get("limit")
        if outstanding is None or limit is None:
            continue
        outstanding_value = float(outstanding)
        limit_value = float(limit)
        status = "YES" if outstanding_value <= limit_value else "NO"
        entries.append(_format_ol_status(status, outstanding_value, limit_value))
        if section_outstanding is None and section_limit is None:
            section_outstanding = outstanding_value
            section_limit = limit_value

    if not entries:
        return "N/A", None, None
    return " | ".join(entries), section_outstanding, section_limit


def score_to_equivalent(score: Optional[int]) -> Optional[str]:
    """Convert credit score to equivalent grade."""
    if score is None:
        return None
    for lower, upper, grade in SCORE_RANGE_EQUIVALENTS:
        if lower <= score <= upper:
            return grade
    return None


def _format_non_bank_mia(stats: Dict[str, Any]) -> Optional[str]:
    """Format non-bank lender MIA counts for display."""
    if not isinstance(stats, dict):
        return None
    
    last_1 = _get_freq(stats, "last_1_month")
    last_6 = _get_freq(stats, "last_6_months")
    
    if not isinstance(last_1, dict) and not isinstance(last_6, dict):
        return None

    def has_any_mia(counts_dict: Dict[str, Any], plus_key: str) -> bool:
        """Return True when any MIA bucket has a non-zero value."""
        return any(_safe_int(counts_dict.get(k, 0)) > 0 for k in ("1", "2", "3", plus_key))

    if (
        (not isinstance(last_6, dict) or not has_any_mia(last_6, "4+"))
        and (not isinstance(last_1, dict) or not has_any_mia(last_1, "4+"))
    ):
        return None

    parts = []
    if isinstance(last_6, dict):
        parts.append(f"past 6 months {_format_mia_bucket_line(last_6, '4+')}")
    if isinstance(last_1, dict):
        parts.append(f"current 1 month {_format_mia_bucket_line(last_1, '4+')}")

    return " and /or ".join(parts) if parts else None


def _get_non_bank_data(non_bank: Dict[str, Any]) -> tuple:
    """Extract non-bank lender data."""
    totals = _safe_get(non_bank, "totals", {})
    stats = _safe_get(non_bank, "stats_totals", {})
    records = _safe_get(non_bank, "records", [])
    
    # Format MIA counts for conduct count display
    mia_formatted = _format_non_bank_mia(stats)
    
    # Conduct count - use formatted MIA string if available, otherwise count records
    conduct_count = mia_formatted
    if conduct_count is None and records:
        # Fallback: count records (accounts) with ANY MIA (1, 2, 3, or 4+) in last 6 months
        count = 0
        for record in records:
            record_stats = record.get("stats", {})
            last_6_values = record_stats.get("last_6_months", {}).get("values", [])
            # Check if this record has ANY MIA (1, 2, 3, or 4+) in the last 6 months
            if any(val is not None and val >= 1 for val in last_6_values):
                count += 1
        conduct_count = count if count > 0 else "N/A"
    elif conduct_count is None:
        conduct_count = "N/A"
    
    # Legal status
    markers = sorted({record.get("legal_marker") for record in records if record.get("legal_marker")})
    legal_status = ", ".join(markers) if markers else "No"
    
    return totals, stats, conduct_count, legal_status


def _within_limit(outstanding, limit) -> str:
    """Check if outstanding is within limit."""
    return "YES" if outstanding is not None and limit is not None and outstanding <= limit else "NO"


def _extract_ccris_legal_status(sections: List[Dict[str, Any]]) -> str:
    """Extract and format CCRIS legal status codes from detailed banking sections."""
    legal_status_details: List[str] = []
    seen: set[str] = set()

    for section in sections:
        details = section.get("account_line_analysis", {}).get("legal_status_details", [])
        if not isinstance(details, list):
            continue
        for detail in details:
            detail_text = str(detail).strip()
            if detail_text and detail_text not in seen:
                seen.add(detail_text)
                legal_status_details.append(detail_text)

    return ", ".join(legal_status_details) if legal_status_details else "N/A"


def _should_highlight_ccris_legal_status(label: str, value: Any) -> bool:
    if not _norm(label).startswith(_norm("CCRIS Loan Account - Legal Status (per primary CRA report)")):
        return False
    if value is None:
        return False
    return str(value).strip().upper() != "N/A"


def _place(out: List[KnockoutCellPlacement], label: str, col_offset: int, value: Any) -> None:
    out.append(KnockoutCellPlacement(label=label, col_offset=col_offset, value=value))


def _place_per_subject(
    out: List[KnockoutCellPlacement],
    label: str,
    num_subjects: int,
    value_fn: Callable[[int], Any],
) -> None:
    for i in range(1, num_subjects + 1):
        _place(out, label, _subject_col_offset(i), value_fn(i))


def build_knockout_placements(merged: Dict[str, Any]) -> List[KnockoutCellPlacement]:
    """Map merged extract JSON to explicit (row label, column offset, value) placements."""
    summary = merged.get("summary_report", {})
    detailed = merged.get("detailed_credit_report", {})
    non_bank = merged.get("non_bank_lender_credit_information", {})

    totals = detailed.get("totals", {})
    total_limit = totals.get("total_limit") or summary.get("Borrower_Total_Limit_RM")
    total_outstanding = totals.get("total_outstanding_balance") or summary.get("Borrower_Outstanding_RM")

    non_bank_totals, _non_bank_stats, non_bank_conduct, non_bank_legal = _get_non_bank_data(non_bank)

    num_subjects = 1
    while summary.get(f"Name_Of_Subject_{num_subjects + 1}"):
        num_subjects += 1

    def get_subject_field(field_name: str, subject_idx: int) -> Any:
        return summary.get(field_name if subject_idx == 1 else f"{field_name}_{subject_idx}")

    placements: List[KnockoutCellPlacement] = []

    _place_per_subject(
        placements,
        LBL_SCORE_RAW,
        num_subjects,
        lambda i: _format_number(get_subject_field("i_SCORE", i)),
    )
    _place_per_subject(
        placements,
        LBL_SCORE_EQ,
        num_subjects,
        lambda i: score_to_equivalent(get_subject_field("i_SCORE", i)),
    )

    _place(placements, LBL_OPS_YEARS, 0, _format_number(summary.get("Incorporation_Year")))
    _place(placements, LBL_COMPANY_STATUS, 0, summary.get("Status"))
    _place(placements, LBL_EXEMPT_PRIVATE, 0, summary.get("Private_Exempt_Company"))

    _place_per_subject(
        placements,
        LBL_WINDING_UP,
        num_subjects,
        lambda i: _format_number(get_subject_field("Winding_Up_Record", i)),
    )
    _place_per_subject(
        placements,
        LBL_CREDIT_APPR,
        num_subjects,
        lambda i: _format_number(get_subject_field("Credit_Applications_Approved_Last_12_months", i)),
    )
    _place_per_subject(
        placements,
        LBL_CREDIT_PEND,
        num_subjects,
        lambda i: _format_number(get_subject_field("Credit_Applications_Pending", i)),
    )
    _place_per_subject(
        placements,
        LBL_LEGAL_ACTION,
        num_subjects,
        lambda i: _format_number(get_subject_field("Legal_Action_taken_from_Banking", i)),
    )
    _place_per_subject(
        placements,
        LBL_EXISTING_FAC,
        num_subjects,
        lambda i: _format_number(get_subject_field("Existing_No_of_Facility_from_Banking", i)),
    )
    _place_per_subject(
        placements,
        LBL_LEGAL_SUITS,
        num_subjects,
        lambda i: _format_number(get_subject_field("Legal_Suits", i)),
    )
    _place_per_subject(
        placements,
        LBL_TRADE_CREDIT,
        num_subjects,
        lambda i: _format_number_or_na(get_subject_field("Trade_Credit_Reference", i)),
    )
    _place_per_subject(
        placements,
        LBL_LEGAL_CASE_STATUS,
        num_subjects,
        lambda i: ", ".join([
            str(get_subject_field("Legal_Suits_Subject_As_Defendant_Defendant_Name", i) or "No"),
            str(get_subject_field("Other_Known_Legal_Suits_Subject_As_Defendant_Defendant_Name", i) or "No"),
            str(get_subject_field("Case_Withdrawn_Settled_Defendant_Name", i) or "No"),
        ]),
    )
    _place_per_subject(
        placements,
        LBL_TOTAL_ENQ,
        num_subjects,
        lambda i: _format_number(get_subject_field("Total_Enquiries_Last_12_months", i)),
    )
    _place_per_subject(
        placements,
        LBL_SPECIAL_ATTN,
        num_subjects,
        lambda i: _format_number(get_subject_field("Special_Attention_Account", i)),
    )

    sections = detailed.get("sections", [])
    if not sections:
        sections = [{"account_line_analysis": detailed.get("account_line_analysis", {})}]

    merged_overdraft = (
        _compute_overdraft_compliance(_merge_overdraft_comparisons(sections)) if sections else "N/A"
    )
    fallback_banking_status = _format_ol_status(
        _within_limit(total_outstanding, total_limit),
        total_outstanding,
        total_limit,
    )
    banking_status_by_section: List[str] = []
    banking_outstanding_by_section: List[Optional[float]] = []
    banking_limit_by_section: List[Optional[float]] = []
    per_section_overdraft: List[str] = []

    for section in sections:
        analysis = section.get("account_line_analysis", {})
        section_status, section_outstanding, section_limit = _compute_banking_facility_status(analysis)
        banking_status_by_section.append(
            section_status if section_status != "N/A" else fallback_banking_status
        )
        banking_outstanding_by_section.append(
            section_outstanding if section_outstanding is not None else total_outstanding
        )
        banking_limit_by_section.append(
            section_limit if section_limit is not None else total_limit
        )
        od = _compute_overdraft_compliance(analysis)
        per_section_overdraft.append(od if od else "N/A")

    if not banking_status_by_section:
        banking_status_by_section = [fallback_banking_status]
    if not banking_outstanding_by_section:
        banking_outstanding_by_section = [total_outstanding]
    if not banking_limit_by_section:
        banking_limit_by_section = [total_limit]
    if not per_section_overdraft:
        per_section_overdraft = [merged_overdraft]

    non_bank_within = _within_limit(non_bank_totals.get("total_outstanding"), non_bank_totals.get("total_limit"))
    ccris_legal_status = _extract_ccris_legal_status(sections)

    for i in range(1, num_subjects + 1):
        col = _subject_col_offset(i)
        sec_i = i - 1
        banking_status = (
            banking_status_by_section[sec_i]
            if sec_i < len(banking_status_by_section)
            else fallback_banking_status
        )
        section_outstanding = (
            banking_outstanding_by_section[sec_i]
            if sec_i < len(banking_outstanding_by_section)
            else total_outstanding
        )
        section_limit = (
            banking_limit_by_section[sec_i]
            if sec_i < len(banking_limit_by_section)
            else total_limit
        )
        overdraft_val = (
            per_section_overdraft[sec_i]
            if sec_i < len(per_section_overdraft)
            else merged_overdraft
        )
        conduct_raw = (
            sections[sec_i].get("account_line_analysis", {}).get("digit_counts_totals")
            if sec_i < len(sections)
            else None
        )

        _place(placements, LBL_OVERDRAFT, col, overdraft_val)
        _place(placements, LBL_BANKING_WITHIN, col, banking_status)
        _place(placements, LBL_SUM_OUT, col, _format_number(section_outstanding))
        _place(placements, LBL_SUM_LIM, col, _format_number(section_limit))
        _place(placements, LBL_NONBANK_WITHIN, col, non_bank_within)
        _place(placements, LBL_CCRIS_CONDUCT, col, conduct_raw)
        _place(placements, LBL_CCRIS_LEGAL, col, ccris_legal_status)
        _place(placements, LBL_NLCI_CONDUCT, col, non_bank_conduct)
        _place(placements, LBL_NLCI_LEGAL, col, non_bank_legal)

    _place(placements, LBL_TOTAL_LIMIT, 0, _format_number(total_limit))
    _place(placements, LBL_TOTAL_OUTSTANDING, 0, _format_number(total_outstanding))

    return placements


def find_issuer_data_column(ws: Worksheet) -> int:
    """Find the 'Issuer' column for Knock-Out Items section."""
    for r in range(1, 11):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and _norm(v) == "issuer":
                return c
    raise ValueError("Cannot find Issuer header column.")


def build_label_row_index(ws: Worksheet, label_col: int = LABEL_COL) -> Dict[str, int]:
    """Map each Knock-Out label text in column D to row number."""
    return {
        _norm(v): r
        for r in range(1, ws.max_row + 1)
        if (v := ws.cell(r, label_col).value) and isinstance(v, str) and v.strip()
    }


def set_issuer_name(ws: Worksheet, issuer_col: int, issuer_name: str) -> None:
    """Set Issuer Name next to 'Issuer Name:' label."""
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 4).value
        if isinstance(v, str) and "issuer name" in _norm(v):
            ws.cell(r, issuer_col).value = issuer_name
            return


def set_cra_report_dates(ws: Worksheet, cra_report_date: Optional[str]) -> None:
    """Set CRA report dates in the worksheet."""
    if not cra_report_date:
        return
    
    for r in range(1, 15):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and _norm(v) == "date (cra report):":
                target_col = c + 1
                for offset in range(1, 4):
                    next_cell = ws.cell(r, c + offset)
                    if isinstance(next_cell.value, str) and "dd/mm/yyyy" in next_cell.value.lower():
                        target_col = c + offset
                        break
                ws.cell(r, target_col).value = cra_report_date
                return


def write_credit_assessment_sheet(wb: openpyxl.Workbook, assessments: List[Assessment]) -> None:
    """Add (or replace) a 'Credit Assessment' sheet with structured assessment output."""
    from openpyxl.styles import Font, PatternFill, Alignment

    SHEET = "Credit Assessment"
    if SHEET in wb.sheetnames:
        del wb[SHEET]
    ws = wb.create_sheet(SHEET)

    BOLD = Font(bold=True)
    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    RED_FILL    = PatternFill("solid", fgColor="C00000")
    RED_FONT    = Font(bold=True, color="FFFFFF")
    ORANGE_FILL = PatternFill("solid", fgColor="FF6600")
    ORANGE_FONT = Font(bold=True, color="FFFFFF")
    BLUE_FILL   = PatternFill("solid", fgColor="2E75B6")
    BLUE_FONT   = Font(bold=True, color="FFFFFF")
    GREEN_FILL  = PatternFill("solid", fgColor="375623")
    GREEN_FONT  = Font(bold=True, color="FFFFFF")

    def _hdr(row: int, col: int, text: str) -> None:
        c = ws.cell(row, col, text)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(wrap_text=True)

    def _label(row: int, col: int, text: str) -> None:
        c = ws.cell(row, col, text)
        c.font = BOLD

    def _val(row: int, col: int, text: Any) -> None:
        ws.cell(row, col, text)

    row = 1
    for a in assessments:
        # ── Header banner ──────────────────────────────────────────
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        banner = ws.cell(row, 1, f"AIgent Credit — Assessment Report  |  {a.company_name}  |  Subject {a.si}  |  {a.date}")
        banner.font = HEADER_FONT
        banner.fill = HEADER_FILL
        banner.alignment = Alignment(horizontal="center")
        row += 1

        # ── Decision summary ───────────────────────────────────────
        _label(row, 1, "Decision"); _val(row, 2, a.recommendation); row += 1
        _label(row, 1, "Risk Band"); _val(row, 2, a.risk_band); row += 1
        _label(row, 1, "Risk Score"); _val(row, 2, f"{a.risk_score} / 100"); row += 1
        _label(row, 1, "CRA Grade"); _val(row, 2, f"{a.grade}  (i-SCORE: {a.cra_score or 'N/A'})"); row += 1
        if a.utilization is not None:
            _label(row, 1, "Credit Utilization")
            _val(row, 2, f"{a.utilization * 100:.1f}%  (RM {a.total_outstanding:,.0f} outstanding / RM {a.total_limit:,.0f} limit)")
        else:
            _label(row, 1, "Credit Utilization"); _val(row, 2, "N/A")
        row += 1
        _label(row, 1, "Years in Operation"); _val(row, 2, a.ops_years if a.ops_years is not None else "N/A"); row += 1
        rec_limit = f"RM {a.limit:,}" if a.recommendation != "DECLINE" else "DECLINE — do not extend credit"
        _label(row, 1, "Recommended Limit"); _val(row, 2, rec_limit); row += 1
        row += 1

        # ── Hard declines ──────────────────────────────────────────
        if a.decline_reasons:
            _hdr(row, 1, "HARD DECLINE TRIGGERS"); row += 1
            for reason in a.decline_reasons:
                c = ws.cell(row, 1, f"✖  {reason}")
                c.font = Font(bold=True, color="C00000")
                row += 1
            row += 1

        # ── Scoring breakdown ──────────────────────────────────────
        _hdr(row, 1, "Dimension"); _hdr(row, 2, "Score"); _hdr(row, 3, "Notes"); row += 1
        for d in a.dimensions:
            _val(row, 1, d.name)
            _val(row, 2, f"{d.score}/{d.max_score}")
            _val(row, 3, "; ".join(d.notes))
            row += 1
        row += 1

        # ── Early warnings ─────────────────────────────────────────
        if a.warnings:
            _hdr(row, 1, "Level"); _hdr(row, 2, "Code"); _hdr(row, 3, "Details"); row += 1
            level_styles = {
                "RED_FLAG": (RED_FILL,    RED_FONT),
                "WARNING":  (ORANGE_FILL, ORANGE_FONT),
                "WATCH":    (BLUE_FILL,   BLUE_FONT),
            }
            for w in a.warnings:
                fill, font = level_styles.get(w.level, (None, BOLD))
                c1 = ws.cell(row, 1, w.level)
                c2 = ws.cell(row, 2, w.code)
                c3 = ws.cell(row, 3, w.message)
                c3.alignment = Alignment(wrap_text=True)
                if fill:
                    c1.fill = fill; c1.font = font
                    c2.fill = fill; c2.font = font
                row += 1
        else:
            c = ws.cell(row, 1, "No early warning signals detected.")
            c.font = Font(color="375623")
            row += 1

        row += 2  # gap between subjects

    # ── Column widths ──────────────────────────────────────────────
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 80
    ws.column_dimensions["D"].width = 20


def fill_knockout_matrix(
    file_path: str,
    issuer_name: str,
    placements: Sequence[KnockoutCellPlacement],
    cra_report_date: Optional[str] = None,
    all_subject_names: Optional[list[str]] = None,
    assessments: Optional[List[Assessment]] = None,
) -> str:
    """Fill the knockout matrix Excel template using explicit cell placements."""
    wb = openpyxl.load_workbook(file_path)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{SHEET_NAME}' not found. Found: {wb.sheetnames}")

    ws = wb[SHEET_NAME]
    
    # Set issuer name
    for r in range(1, 35):
        if isinstance(ws.cell(r, 4).value, str) and "issuer name" in _norm(ws.cell(r, 4).value):
            set_issuer_name(ws, 5, issuer_name)
            break
    
    set_cra_report_dates(ws, cra_report_date)
    issuer_data_col = find_issuer_data_column(ws)
    
    # Insert subject names
    if all_subject_names is None:
        all_subject_names = [issuer_name]

    # Always map subject names in +2 column steps (issuer, subject_2, subject_3, ...)
    # so every extracted name gets its own slot like other multi-subject fields.
    subject_cols: list[int] = []
    for i, _ in enumerate(all_subject_names):
        col = issuer_data_col + i * 2
        if col <= ws.max_column:
            subject_cols.append(col)
        else:
            print(
                f"⚠️ Skipping subject name at index {i + 1}: template has no column {col} "
                f"(max column: {ws.max_column})"
            )
            break

    inserted_subject_cols: list[int] = []
    for i, subject_name in enumerate(all_subject_names[:len(subject_cols)]):
        if subject_name:
            col = subject_cols[i]
            ws.cell(7, col).value = subject_name
            inserted_subject_cols.append(col)

    if len(all_subject_names) > len(subject_cols):
        print(
            f"⚠️ Inserted {len(subject_cols)} subject name(s) out of {len(all_subject_names)}; "
            "remaining names were not written because template columns ran out."
        )
    
    label_index = build_label_row_index(ws, LABEL_COL)

    missing: List[str] = []
    written = 0

    for p in placements:
        row = label_index.get(_norm(p.label))
        if not row:
            missing.append(p.label)
            continue
        target_col = issuer_data_col + p.col_offset
        if target_col > ws.max_column:
            print(
                f"⚠️ Skip '{p.label}' at col {target_col}: exceeds max_column {ws.max_column}"
            )
            continue
        formatted_value = _format_cell_value(p.value)
        ws.cell(row, target_col).value = formatted_value
        if _should_highlight_ccris_legal_status(p.label, formatted_value):
            ws.cell(row, target_col).font = RED_BOLD_FONT
        written += 1

    # Apply Column L coloring immediately after insertion, only for columns inserted this run.
    cols_to_color = inserted_subject_cols or subject_cols
    highlighted = apply_column_l_highlighting(ws, cols_to_color)
    print(f"🎨 Column L coloring applied: {highlighted} cell(s) highlighted across {len(cols_to_color)} subject column(s)")

    # Write credit assessment sheet if provided
    if assessments:
        write_credit_assessment_sheet(wb, assessments)
        print(f"📊 Credit Assessment sheet written for {len(assessments)} subject(s)")

    # Save output
    output_path = f"{os.path.splitext(file_path)[0]}_FILLED{os.path.splitext(file_path)[1]}"
    wb.save(output_path)

    if missing:
        print("⚠️ Missing labels:")
        for m in dict.fromkeys(missing):
            print(f"  - {m}")

    return output_path


def _find_excel_template(excel_name: str = DEFAULT_EXCEL) -> Optional[str]:
    """
    Find the Excel template file. Checks multiple locations:
    1. PyInstaller temp directory (if bundled)
    2. Same directory as EXE/script
    3. Current working directory
    """
    search_paths = []
    
    # If running as EXE, check PyInstaller's temp directory first
    if getattr(sys, 'frozen', False):
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        if hasattr(sys, '_MEIPASS'):
            search_paths.append(Path(sys._MEIPASS))
        # Also check EXE directory
        search_paths.append(Path(sys.executable).parent)
    
    # Always check current working directory
    search_paths.append(Path.cwd())
    
    # If running as script, also check script directory
    if not getattr(sys, 'frozen', False):
        search_paths.append(Path(__file__).resolve().parent)
    
    # Search all paths
    for path in search_paths:
        excel_path = path / excel_name
        if excel_path.exists():
            return str(excel_path)
    
    return None


def main() -> None:
    """Main entry point."""
    try:
        parser = argparse.ArgumentParser(description="Fill Knockout Matrix Excel from merged credit report data.")
        parser.add_argument("--excel", help="Path to Knockout Matrix Template.xlsx (defaults to template in same directory)")
        parser.add_argument("--merged-json", help="Path to merged JSON output (skips PDF processing)")
        parser.add_argument("--pdf", help="Path to Experian PDF (opens picker if omitted)")
        parser.add_argument("--issuer", help="Issuer name (defaults to Name Of Subject from PDF)")
        args = parser.parse_args()

        # Get Excel file path (works in both script and EXE)
        if args.excel:
            excel_file = args.excel
            if not os.path.exists(excel_file):
                print(f"❌ Excel template not found at specified path: {excel_file}")
                raise SystemExit(1)
        else:
            # Try to find the Excel template automatically
            excel_file = _find_excel_template()
            if not excel_file:
                print(f"❌ Excel template '{DEFAULT_EXCEL}' not found!")
                print(f"\n💡 Please ensure '{DEFAULT_EXCEL}' is in one of these locations:")
                if getattr(sys, 'frozen', False):
                    print(f"   1. Same folder as the EXE: {Path(sys.executable).parent}")
                    print(f"   2. Current working directory: {Path.cwd()}")
                else:
                    print(f"   1. Same folder as the script: {Path(__file__).resolve().parent}")
                    print(f"   2. Current working directory: {Path.cwd()}")
                print(f"\n   OR use --excel argument to specify the full path")
                raise SystemExit(1)
            print(f"📄 Found Excel template: {excel_file}")
        
        # Load or generate merged report
        if args.merged_json:
            print("📄 Loading merged report from JSON...")
            if not os.path.exists(args.merged_json):
                print(f"❌ Merged JSON file not found: {args.merged_json}")
                raise SystemExit(1)
            with open(args.merged_json, "r", encoding="utf-8") as f:
                merged = json.load(f)
            print("✅ Merged report loaded")
        else:
            # Get PDF path (opens picker if not provided)
            pdf_path = resolve_pdf_path(args.pdf)
            if not pdf_path:
                print("❌ No PDF file selected")
                raise SystemExit(1)
            
            print(f"📄 Processing PDF: {os.path.basename(pdf_path)}")
            print("📊 Generating merged report (this may take a moment for large PDFs)...")
            print("💡 Tip: Save merged JSON with 'python merged_credit_report.py --pdf file.pdf' for faster subsequent runs")
            merged = merge_reports(pdf_path)

        summary = merged.get("summary_report", {})
        issuer_name = args.issuer or summary.get("Name_Of_Subject") or "UNKNOWN ISSUER"

        # Keep all detected subject names (issuer + directors/guarantors) so they can
        # be written to the subject header columns in the knockout template.
        raw_subject_names = summary.get("all_names_of_subject") or []
        all_subject_names = [
            re.sub(r"\s+", " ", str(name)).strip()
            for name in raw_subject_names
            if name and str(name).strip()
        ]

        # Ensure issuer is always the first displayed subject.
        if issuer_name and issuer_name not in all_subject_names:
            all_subject_names.insert(0, issuer_name)
        
        placements = build_knockout_placements(merged)

        # Run credit assessment for all subjects
        num_subjects = 1
        while summary.get(f"Name_Of_Subject_{num_subjects + 1}"):
            num_subjects += 1
        print(f"\n🔍 Running credit assessment for {num_subjects} subject(s)...")
        assessments: List[Assessment] = []
        for si in range(1, num_subjects + 1):
            a = assess(merged, si)
            assessments.append(a)
            decision_icon = "✅" if a.recommendation == "APPROVE" else ("⚠️" if a.recommendation == "CONDITIONAL APPROVE" else "❌")
            print(f"  {decision_icon} Subject {si} ({a.company_name}): {a.recommendation}  |  Risk Score {a.risk_score}/100  |  {a.risk_band}")
            if a.recommendation != "DECLINE":
                print(f"     Recommended Limit: RM {a.limit:,}")
            if a.decline_reasons:
                for r in a.decline_reasons:
                    print(f"     ✖ {r}")

        # Fill Excel
        print(f"\n📝 Filling Excel template: {os.path.basename(excel_file)}")
        output = fill_knockout_matrix(
            excel_file,
            issuer_name,
            placements,
            cra_report_date=summary.get("Last_Updated_By_Experian"),
            all_subject_names=all_subject_names or None,
            assessments=assessments,
        )
        print(f"\n✅ Success! File saved: {os.path.basename(output)}")
        print(f"📁 Location: {os.path.dirname(os.path.abspath(output))}")
    
    except KeyboardInterrupt:
        print("\n❌ Operation cancelled by user")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ Error: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
