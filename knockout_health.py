"""Evaluate Knockout Matrix column-L health from merged credit data (no Excel fill required)."""

from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

import openpyxl

from column_l_validator import CRITERIA_COL, LABEL_COL, SHEET_NAME, _matches_criteria
from insert_excel_file import (
    DEFAULT_EXCEL,
    KnockoutCellPlacement,
    _find_excel_template,
    _format_cell_value,
    _subject_col_offset,
    build_knockout_placements,
    find_issuer_data_column,
)
from text_normalize import normalize_compare_text

_CRITERIA_ROW_START = 11


def _norm_label(s: Any) -> str:
    """Match insert_excel_file label normalization for template / placement labels."""
    return normalize_compare_text(s, smart_typography=True)


@dataclass
class KnockoutHit:
    row: int
    label: str
    criteria: str
    value: Any


@dataclass
class KnockoutHealthResult:
    is_healthy: bool
    subject_index: int
    hits: List[KnockoutHit] = field(default_factory=list)
    unresolved_rows: List[Dict[str, Any]] = field(default_factory=list)


def read_knockout_criteria_rows(template_path: str) -> List[Tuple[int, str, str]]:
    """
    Read Knock-Out rows that have a non-empty column L criterion.

    Returns:
        List of (excel_row, label_text, criteria_text).
    """
    path = Path(template_path)
    if not path.is_file():
        raise FileNotFoundError(f"Knockout template not found: {template_path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{SHEET_NAME}' not in workbook: {wb.sheetnames}")

    ws = wb[SHEET_NAME]
    rows: List[Tuple[int, str, str]] = []
    try:
        for row in range(_CRITERIA_ROW_START, ws.max_row + 1):
            label_cell = ws.cell(row, LABEL_COL).value
            crit_cell = ws.cell(row, CRITERIA_COL).value
            if crit_cell is None or not str(crit_cell).strip():
                continue
            label_text = str(label_cell).strip() if label_cell is not None else ""
            rows.append((row, label_text, str(crit_cell).strip()))
    finally:
        wb.close()

    return rows


def formatted_value_from_placements(
    placements: Sequence[KnockoutCellPlacement],
    label: str,
    subject_index: int,
) -> Any:
    """
    Cell value as written by fill_knockout_matrix: match row label + subject column offset.
    """
    if subject_index < 1:
        raise ValueError("subject_index must be >= 1")
    target_off = _subject_col_offset(subject_index)
    nlabel = _norm_label(label)
    for p in placements:
        if p.col_offset != target_off:
            continue
        if _norm_label(p.label) == nlabel:
            return _format_cell_value(p.value)
    return None


def evaluate_knockout_health(
    merged: Dict[str, Any],
    subject_index: int = 1,
    template_path: Optional[str] = None,
    placements: Optional[Sequence[KnockoutCellPlacement]] = None,
) -> KnockoutHealthResult:
    """
    True if no column-L knockout rule matches for this subject.

    subject_index is 1-based (1 = Issuer column, 2 = next subject column, …).
    """
    if subject_index < 1:
        raise ValueError("subject_index must be >= 1")

    tpl = template_path or _find_excel_template(DEFAULT_EXCEL)
    if not tpl:
        raise FileNotFoundError(
            f"Could not locate '{DEFAULT_EXCEL}'. Pass template_path= explicitly."
        )

    pl = list(placements) if placements is not None else build_knockout_placements(merged)
    criteria_rows = read_knockout_criteria_rows(tpl)

    hits: List[KnockoutHit] = []
    unresolved: List[Dict[str, Any]] = []

    for row, label, criteria in criteria_rows:
        value = formatted_value_from_placements(pl, label, subject_index)
        if value is None and label.strip():
            unresolved.append({"row": row, "label": label, "criteria": criteria})

        if _matches_criteria(criteria, value, label):
            hits.append(KnockoutHit(row=row, label=label, criteria=criteria, value=value))

    return KnockoutHealthResult(
        is_healthy=len(hits) == 0,
        subject_index=subject_index,
        hits=hits,
        unresolved_rows=unresolved,
    )


def validate_knockout_health_vs_excel(
    filled_excel_path: str,
    merged: Dict[str, Any],
    subject_index: int = 1,
    placements: Optional[Sequence[KnockoutCellPlacement]] = None,
) -> Dict[str, Any]:
    """
    Compare programmatic knockout evaluation to values already written in a *_FILLED.xlsx.

    Returns agreement on whether each criteria row matches, and summary counts.
    """
    wb = openpyxl.load_workbook(filled_excel_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{SHEET_NAME}' not in {filled_excel_path}")

    ws = wb[SHEET_NAME]
    issuer_col = find_issuer_data_column(ws)
    subject_col = issuer_col + (subject_index - 1) * 2
    if subject_col > ws.max_column:
        wb.close()
        raise ValueError(
            f"subject_index={subject_index} maps to column {subject_col}, past max_column"
        )

    pl = list(placements) if placements is not None else build_knockout_placements(merged)
    criteria_rows = read_knockout_criteria_rows(filled_excel_path)

    mismatches: List[Dict[str, Any]] = []
    prog_hits = 0
    excel_hits = 0

    for row, label, criteria in criteria_rows:
        excel_val = ws.cell(row, subject_col).value
        prog_val = formatted_value_from_placements(pl, label, subject_index)
        excel_match = _matches_criteria(criteria, excel_val, label)
        prog_match = _matches_criteria(criteria, prog_val, label)
        if excel_match:
            excel_hits += 1
        if prog_match:
            prog_hits += 1
        if excel_match != prog_match:
            mismatches.append(
                {
                    "row": row,
                    "label": label,
                    "criteria": criteria,
                    "excel_value": excel_val,
                    "programmatic_value": prog_val,
                    "excel_match": excel_match,
                    "programmatic_match": prog_match,
                }
            )

    wb.close()

    return {
        "filled_excel": str(filled_excel_path),
        "subject_index": subject_index,
        "subject_column": subject_col,
        "criteria_rows_evaluated": len(criteria_rows),
        "excel_knockout_hits": excel_hits,
        "programmatic_knockout_hits": prog_hits,
        "match_agreement": len(mismatches) == 0,
        "mismatches": mismatches,
    }


def _result_to_dict(result: KnockoutHealthResult) -> Dict[str, Any]:
    return {
        "is_healthy": result.is_healthy,
        "subject_index": result.subject_index,
        "knockout_hit_count": len(result.hits),
        "hits": [
            {"row": h.row, "label": h.label, "criteria": h.criteria, "value": h.value}
            for h in result.hits
        ],
        "unresolved_row_count": len(result.unresolved_rows),
        "unresolved_rows": result.unresolved_rows,
    }


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Evaluate Knockout column-L health from merged credit JSON."
    )
    parser.add_argument("--merged-json", required=True, help="Path to merged_credit_report JSON")
    parser.add_argument(
        "--excel",
        help=f"Knockout template path (default: find {DEFAULT_EXCEL})",
    )
    parser.add_argument(
        "--subject",
        type=int,
        default=1,
        help="1-based subject column (1 = Issuer)",
    )
    parser.add_argument(
        "--validate-filled",
        metavar="XLSX",
        help="After evaluation, compare to this *_FILLED.xlsx workbook",
    )
    args = parser.parse_args()

    path = Path(args.merged_json)
    if not path.is_file():
        print(f"❌ File not found: {path}", file=sys.stderr)
        sys.exit(1)

    with open(path, encoding="utf-8") as f:
        merged = json.load(f)

    try:
        result = evaluate_knockout_health(
            merged,
            subject_index=args.subject,
            template_path=args.excel,
        )
    except (FileNotFoundError, ValueError) as e:
        print(f"❌ {e}", file=sys.stderr)
        sys.exit(1)

    out = _result_to_dict(result)
    print(json.dumps(out, indent=2, ensure_ascii=False))

    if args.validate_filled:
        vpath = Path(args.validate_filled)
        if not vpath.is_file():
            print(f"❌ --validate-filled file not found: {vpath}", file=sys.stderr)
            sys.exit(1)
        cmp = validate_knockout_health_vs_excel(str(vpath), merged, args.subject)
        print(json.dumps({"validation_vs_filled": cmp}, indent=2, ensure_ascii=False))
        if not cmp["match_agreement"]:
            sys.exit(2)


if __name__ == "__main__":
    main()
