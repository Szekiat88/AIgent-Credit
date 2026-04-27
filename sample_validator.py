"""
Batch sample validator — AIgent Credit extraction pipeline.

Walks samples/, runs the pipeline on every PDF, compares the generated
_FILLED.xlsx against the human-filled reference Knockout_*.xlsx, and
highlights mismatched cells orange.

Usage:
    python sample_validator.py                    # all cases
    python sample_validator.py --case Halalgel    # one case only
    python sample_validator.py --no-highlight     # compare without colouring
    python sample_validator.py --no-generate      # skip pipeline, compare existing _FILLED files
"""

from __future__ import annotations

import argparse
import re
import shutil
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl
from openpyxl.styles import PatternFill

from insert_excel_file import (
    build_knockout_placements,
    fill_knockout_matrix,
    find_issuer_data_column,
    SHEET_NAME,
    LABEL_COL,
    _find_excel_template,
    _norm,
    LBL_SCORE_RAW, LBL_SCORE_EQ, LBL_OPS_YEARS, LBL_COMPANY_STATUS,
    LBL_EXEMPT_PRIVATE, LBL_WINDING_UP, LBL_CREDIT_APPR, LBL_CREDIT_PEND,
    LBL_LEGAL_ACTION, LBL_EXISTING_FAC, LBL_LEGAL_SUITS, LBL_TRADE_CREDIT,
    LBL_LEGAL_CASE_STATUS, LBL_TOTAL_ENQ, LBL_SPECIAL_ATTN, LBL_OVERDRAFT,
    LBL_BANKING_WITHIN, LBL_SUM_OUT, LBL_SUM_LIM, LBL_NONBANK_WITHIN,
    LBL_CCRIS_CONDUCT, LBL_CCRIS_LEGAL, LBL_NLCI_CONDUCT, LBL_NLCI_LEGAL,
    LBL_TOTAL_LIMIT, LBL_TOTAL_OUTSTANDING,
)

# Only compare rows the pipeline actually fills — avoids false positives on
# header rows, metadata rows, and manually-filled financial/KYC rows.
COMPARABLE_LABELS: frozenset[str] = frozenset(
    _norm(lbl) for lbl in [
        LBL_SCORE_RAW, LBL_SCORE_EQ, LBL_OPS_YEARS, LBL_COMPANY_STATUS,
        LBL_EXEMPT_PRIVATE, LBL_WINDING_UP, LBL_CREDIT_APPR, LBL_CREDIT_PEND,
        LBL_LEGAL_ACTION, LBL_EXISTING_FAC, LBL_LEGAL_SUITS, LBL_TRADE_CREDIT,
        LBL_LEGAL_CASE_STATUS, LBL_TOTAL_ENQ, LBL_SPECIAL_ATTN, LBL_OVERDRAFT,
        LBL_BANKING_WITHIN, LBL_SUM_OUT, LBL_SUM_LIM, LBL_NONBANK_WITHIN,
        LBL_CCRIS_CONDUCT, LBL_CCRIS_LEGAL, LBL_NLCI_CONDUCT, LBL_NLCI_LEGAL,
        LBL_TOTAL_LIMIT, LBL_TOTAL_OUTSTANDING,
    ]
)
from merged_credit_report import merge_reports
from credit_analyst import assess

SAMPLES_DIR  = Path(__file__).resolve().parent / "samples"
OUTPUT_DIR   = Path(__file__).resolve().parent / "samples_output"
ORANGE_FILL  = PatternFill("solid", fgColor="FFA500")


# ─── Data classes ─────────────────────────────────────────────────────────────

@dataclass
class CellDiff:
    row: int
    col: int
    label: str
    generated: str
    reference: str


@dataclass
class PDFResult:
    pdf_name: str
    filled_path: Optional[Path]
    total_compared: int = 0
    diffs: List[CellDiff] = field(default_factory=list)
    error: Optional[str] = None

    @property
    def matches(self) -> int:
        return self.total_compared - len(self.diffs)


@dataclass
class CaseResult:
    case_name: str
    pdfs: List[Path]
    reference_excel: Optional[Path]
    pdf_results: List[PDFResult] = field(default_factory=list)


# ─── Value normalisation ──────────────────────────────────────────────────────

def _normalise(v) -> str:
    """Normalise a cell value for comparison (type-agnostic)."""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%d/%m/%Y")
    s = str(v).strip()
    if not s:
        return ""
    upper = s.upper().replace(" ", "")
    if upper in ("N/A", "NA", "NONE", "-"):
        return "n/a"
    # Numeric normalisation: strip commas/spaces, compare as number
    cleaned = re.sub(r"[\s,]", "", s)
    try:
        f = float(cleaned)
        return str(int(f)) if f == int(f) else f"{f:.4f}".rstrip("0").rstrip(".")
    except (ValueError, OverflowError):
        pass
    return s.lower()


# ─── Case discovery ───────────────────────────────────────────────────────────

def discover_cases(samples_dir: Path, name_filter: Optional[str] = None) -> List[dict]:
    """Return per-case dicts from the samples directory."""
    cases = []
    for folder in sorted(samples_dir.iterdir()):
        if not folder.is_dir() or folder.name.startswith("."):
            continue
        if name_filter and folder.name.lower() != name_filter.lower():
            continue
        pdfs = sorted(folder.glob("*.pdf"))
        refs = sorted(folder.glob("Knockout_*.xlsx"))
        if not pdfs:
            continue
        cases.append({
            "name": folder.name,
            "path": folder,
            "pdfs": pdfs,
            "reference": refs[0] if refs else None,
        })
    return cases


# ─── Pipeline runner ──────────────────────────────────────────────────────────

def run_pipeline(pdf_path: Path, template_path: Path, output_folder: Path) -> Path:
    """
    Copy the template into output_folder, fill it from the PDF, return the
    _FILLED path. Cleans up the template copy afterward.
    output_folder is always under samples_output/ so the read-only samples/
    tree is never written to.
    """
    output_folder.mkdir(parents=True, exist_ok=True)
    working_copy = output_folder / f"{pdf_path.stem}.xlsx"
    shutil.copy2(str(template_path), str(working_copy))

    try:
        merged = merge_reports(str(pdf_path))
        summary = merged.get("summary_report", {})
        issuer_name = summary.get("Name_Of_Subject") or "UNKNOWN ISSUER"

        raw_names = summary.get("all_names_of_subject") or []
        all_subject_names = [
            re.sub(r"\s+", " ", str(n)).strip()
            for n in raw_names if n and str(n).strip()
        ]
        if issuer_name and issuer_name not in all_subject_names:
            all_subject_names.insert(0, issuer_name)

        placements = build_knockout_placements(merged)

        num_subjects = 1
        while summary.get(f"Name_Of_Subject_{num_subjects + 1}"):
            num_subjects += 1
        assessments = [assess(merged, si) for si in range(1, num_subjects + 1)]

        filled_path = fill_knockout_matrix(
            str(working_copy),
            issuer_name,
            placements,
            cra_report_date=summary.get("Last_Updated_By_Experian"),
            all_subject_names=all_subject_names or None,
            assessments=assessments,
        )
        return Path(filled_path)
    finally:
        if working_copy.exists():
            working_copy.unlink()


# ─── Cell comparison ─────────────────────────────────────────────────────────

def _detect_label_col(ws) -> int:
    """Detect which column holds row labels (usually D=4, but older templates use E=5)."""
    # Try columns 4 and 5; pick whichever has more COMPARABLE_LABELS matches.
    best_col, best_count = LABEL_COL, 0
    for c in (4, 5):
        count = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and _norm(v) in COMPARABLE_LABELS:
                count += 1
        if count > best_count:
            best_col, best_count = c, count
    return best_col


def compare_sheets(gen_ws, ref_ws) -> tuple[int, List[CellDiff]]:
    """
    Compare data-zone cells between generated and reference Knock-Out sheets.
    Data zone: rows with a label in column D (or E for older templates), columns
    from issuer_col rightward.  Returns (total_cells_compared, list_of_diffs).
    """
    try:
        gen_issuer_col = find_issuer_data_column(gen_ws)
    except ValueError:
        return 0, []

    try:
        ref_issuer_col = find_issuer_data_column(ref_ws)
    except ValueError:
        return 0, []

    ref_label_col = _detect_label_col(ref_ws)

    # Build a label→row map for the reference sheet so we can align by label,
    # not by row number (handles minor row-count differences between versions).
    ref_label_to_row: Dict[str, int] = {}
    for r in range(1, ref_ws.max_row + 1):
        v = ref_ws.cell(r, ref_label_col).value
        if isinstance(v, str) and v.strip():
            key = v.strip().lower()[:80]
            ref_label_to_row[key] = r

    # Cap comparison at the rightmost subject-name column in row 7
    # (avoids secondary template lookup blocks further to the right).
    def _last_data_col(ws, from_col: int) -> int:
        last = from_col
        for c in range(from_col, ws.max_column + 1):
            if ws.cell(7, c).value is not None:
                last = c
        return last

    max_gen_col = _last_data_col(gen_ws, gen_issuer_col)
    max_ref_col = _last_data_col(ref_ws, ref_issuer_col)

    total = 0
    diffs: List[CellDiff] = []

    for gen_row in range(1, gen_ws.max_row + 1):
        label_val = gen_ws.cell(gen_row, LABEL_COL).value
        if not isinstance(label_val, str) or not label_val.strip():
            continue

        # Only compare rows the pipeline fills — skip header/metadata/KYC rows
        if _norm(label_val) not in COMPARABLE_LABELS:
            continue

        label_key = label_val.strip().lower()[:80]
        ref_row = ref_label_to_row.get(label_key)
        if ref_row is None:
            continue  # label not in reference → skip

        # Compare columns: use relative offset from each sheet's own issuer col
        max_offset = min(max_gen_col - gen_issuer_col, max_ref_col - ref_issuer_col)
        for offset in range(max_offset + 1):
            gen_col = gen_issuer_col + offset
            ref_col = ref_issuer_col + offset

            gen_val = gen_ws.cell(gen_row, gen_col).value
            ref_val = ref_ws.cell(ref_row, ref_col).value

            gen_norm = _normalise(gen_val)
            ref_norm = _normalise(ref_val)

            # Skip cells where both sides are empty
            if not gen_norm and not ref_norm:
                continue

            total += 1
            if gen_norm != ref_norm:
                diffs.append(CellDiff(
                    row=gen_row, col=gen_col,
                    label=label_val.strip()[:60],
                    generated=str(gen_val) if gen_val is not None else "",
                    reference=str(ref_val) if ref_val is not None else "",
                ))

    return total, diffs


def compare_excels(generated_path: Path, reference_path: Path) -> tuple[int, List[CellDiff]]:
    """Load both workbooks and compare their Knock-Out sheets."""
    gen_wb = openpyxl.load_workbook(str(generated_path), data_only=True)
    ref_wb = openpyxl.load_workbook(str(reference_path), data_only=True)

    if SHEET_NAME not in gen_wb.sheetnames or SHEET_NAME not in ref_wb.sheetnames:
        return 0, []

    return compare_sheets(gen_wb[SHEET_NAME], ref_wb[SHEET_NAME])


# ─── Highlight mismatches ─────────────────────────────────────────────────────

def highlight_mismatches(filled_path: Path, diffs: List[CellDiff]) -> None:
    """Apply orange fill to every mismatched cell in the filled Excel."""
    if not diffs:
        return
    wb = openpyxl.load_workbook(str(filled_path))
    if SHEET_NAME not in wb.sheetnames:
        return
    ws = wb[SHEET_NAME]
    for d in diffs:
        ws.cell(d.row, d.col).fill = ORANGE_FILL
    wb.save(str(filled_path))


# ─── Summary printer ─────────────────────────────────────────────────────────

def print_summary(results: List[CaseResult]) -> None:
    W = 90
    print()
    print("═" * W)
    print("  AIgent Credit — Sample Validation Report")
    print("═" * W)
    print(f"  {'Case':<28} {'PDF':<36} {'Ref':>3}  {'✓':>5}  {'✗':>5}  Status")
    print("─" * W)

    grand_match = grand_miss = 0

    for r in results:
        ref_mark = "✓" if r.reference_excel else "✗"

        if not r.pdf_results:
            print(f"  {r.case_name:<28} {'(no results)':<36} {ref_mark:>3}  {'—':>5}  {'—':>5}  ⚠ SKIPPED")
            continue

        for pr in r.pdf_results:
            name_trunc = r.case_name[:28]
            pdf_trunc = pr.pdf_name[:36]

            if pr.error:
                print(f"  {name_trunc:<28} {pdf_trunc:<36} {ref_mark:>3}  {'—':>5}  {'—':>5}  ❌ ERROR: {pr.error[:30]}")
                continue

            if not r.reference_excel:
                print(f"  {name_trunc:<28} {pdf_trunc:<36} {ref_mark:>3}  {'—':>5}  {'—':>5}  ✓ GENERATED ONLY")
                continue

            m, x = pr.matches, len(pr.diffs)
            grand_match += m
            grand_miss  += x
            status = "✅ PASS" if x == 0 else f"⚠ {x} MISMATCH(ES)"
            print(f"  {name_trunc:<28} {pdf_trunc:<36} {ref_mark:>3}  {m:>5}  {x:>5}  {status}")

            # Print first 5 diffs inline
            for diff in pr.diffs[:5]:
                lbl = diff.label[:45]
                print(f"      ↳ row {diff.row} col {diff.col}  [{lbl}]")
                print(f"           GEN: {repr(diff.generated[:60])}")
                print(f"           REF: {repr(diff.reference[:60])}")
            if len(pr.diffs) > 5:
                print(f"      … and {len(pr.diffs) - 5} more (see orange cells in _FILLED.xlsx)")

    print("─" * W)
    print(f"  Grand total  Matches: {grand_match}   Mismatches: {grand_miss}")
    print("═" * W)
    print()


# ─── Main ─────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Validate AIgent Credit pipeline output against sample reference Excels."
    )
    parser.add_argument("--case", help="Run only the named case folder (case-insensitive)")
    parser.add_argument("--no-highlight", action="store_true",
                        help="Skip orange highlighting (compare only)")
    parser.add_argument("--no-generate", action="store_true",
                        help="Skip pipeline run; compare existing _FILLED.xlsx files instead")
    args = parser.parse_args()

    if not SAMPLES_DIR.exists():
        print(f"❌ samples/ directory not found at: {SAMPLES_DIR}")
        sys.exit(1)

    template_path = _find_excel_template()
    if not template_path and not args.no_generate:
        print("❌ Knockout Matrix Template.xlsx not found — cannot generate filled Excels.")
        sys.exit(1)

    cases = discover_cases(SAMPLES_DIR, name_filter=args.case)
    if not cases:
        target = args.case or "any subfolder"
        print(f"❌ No cases found matching: {target}")
        sys.exit(1)

    results: List[CaseResult] = []

    for case in cases:
        print(f"\n{'─'*60}")
        print(f"  Case: {case['name']}  ({len(case['pdfs'])} PDF(s))")
        print(f"{'─'*60}")

        cr = CaseResult(
            case_name=case["name"],
            pdfs=case["pdfs"],
            reference_excel=case["reference"],
        )

        for pdf_path in case["pdfs"]:
            print(f"  📄 {pdf_path.name}")
            pr = PDFResult(pdf_name=pdf_path.name, filled_path=None)

            # ── Generate ──────────────────────────────────────────
            out_folder = OUTPUT_DIR / case["name"]

            if args.no_generate:
                candidate = out_folder / f"{pdf_path.stem}_FILLED.xlsx"
                if candidate.exists():
                    pr.filled_path = candidate
                    print(f"     Using existing: {candidate.name}")
                else:
                    pr.error = f"_FILLED.xlsx not found ({candidate.name})"
                    print(f"     ⚠ {pr.error}")
                    cr.pdf_results.append(pr)
                    continue
            else:
                try:
                    print(f"     Running pipeline…", end="", flush=True)
                    pr.filled_path = run_pipeline(pdf_path, Path(template_path), out_folder)
                    print(f" ✓  → {pr.filled_path.name}")
                except Exception as exc:
                    pr.error = str(exc)
                    print(f"\n     ❌ Pipeline error: {exc}")
                    cr.pdf_results.append(pr)
                    continue

            # ── Compare ───────────────────────────────────────────
            if case["reference"] and pr.filled_path:
                try:
                    total, diffs = compare_excels(pr.filled_path, case["reference"])
                    pr.total_compared = total
                    pr.diffs = diffs
                    match_count = total - len(diffs)
                    print(f"     Compared {total} cell(s): {match_count} match, {len(diffs)} mismatch")
                except Exception as exc:
                    pr.error = f"Comparison failed: {exc}"
                    print(f"     ❌ {pr.error}")

                # ── Highlight ─────────────────────────────────────
                if pr.diffs and not args.no_highlight and pr.filled_path:
                    try:
                        highlight_mismatches(pr.filled_path, pr.diffs)
                        print(f"     🟠 {len(pr.diffs)} cell(s) highlighted orange in {pr.filled_path.name}")
                    except Exception as exc:
                        print(f"     ⚠ Could not highlight: {exc}")
            else:
                print(f"     ℹ No reference Excel — skipping comparison")

            cr.pdf_results.append(pr)

        results.append(cr)

    print_summary(results)


if __name__ == "__main__":
    main()
