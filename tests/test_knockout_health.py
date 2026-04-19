"""Tests for knockout_health (minimal template + merged JSON)."""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

import openpyxl

from column_l_validator import CRITERIA_COL, SHEET_NAME
from insert_excel_file import LABEL_COL, LBL_OPS_YEARS
from knockout_health import (
    evaluate_knockout_health,
    formatted_value_from_placements,
    validate_knockout_health_vs_excel,
)
from insert_excel_file import build_knockout_placements


def _minimal_merged(incorporation_year: int) -> dict:
    return {
        "summary_report": {
            "Incorporation_Year": incorporation_year,
            "Name_Of_Subject": "Test Co",
        },
        "detailed_credit_report": {},
        "non_bank_lender_credit_information": {},
    }


def _write_min_knockout_template(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    # Issuer header (find_issuer_data_column scans rows 1–10)
    ws.cell(1, 13).value = "Issuer"
    ws.cell(11, LABEL_COL).value = LBL_OPS_YEARS
    ws.cell(11, CRITERIA_COL).value = "no"
    wb.save(path)
    wb.close()


class KnockoutHealthTests(unittest.TestCase):
    def test_formatted_value_from_placements_miss(self) -> None:
        pl = build_knockout_placements(_minimal_merged(5))
        v = formatted_value_from_placements(pl, "Nonexistent Label", 1)
        self.assertIsNone(v)

    def test_operations_years_knockout(self) -> None:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tpath = Path(tmp.name)
        try:
            _write_min_knockout_template(tpath)
            bad = evaluate_knockout_health(
                _minimal_merged(2), subject_index=1, template_path=str(tpath)
            )
            self.assertFalse(bad.is_healthy)
            self.assertEqual(len(bad.hits), 1)

            good = evaluate_knockout_health(
                _minimal_merged(5), subject_index=1, template_path=str(tpath)
            )
            self.assertTrue(good.is_healthy)
            self.assertEqual(len(good.hits), 0)
        finally:
            tpath.unlink(missing_ok=True)

    def test_validate_vs_excel_agrees(self) -> None:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tpath = Path(tmp.name)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp2:
            fpath = Path(tmp2.name)
        try:
            _write_min_knockout_template(tpath)
            merged = _minimal_merged(5)
            # Filled workbook: copy template structure and write the cell Excel would hold
            wb = openpyxl.load_workbook(tpath)
            ws = wb[SHEET_NAME]
            ws.cell(11, 13).value = "5"
            wb.save(fpath)
            wb.close()

            rep = validate_knockout_health_vs_excel(str(fpath), merged, subject_index=1)
            self.assertTrue(rep["match_agreement"])
            self.assertEqual(rep["excel_knockout_hits"], rep["programmatic_knockout_hits"])
        finally:
            tpath.unlink(missing_ok=True)
            fpath.unlink(missing_ok=True)


if __name__ == "__main__":
    unittest.main()
