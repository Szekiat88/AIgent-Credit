"""Validate inserted Knock-Out values against Column L criteria and highlight matches."""

from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

SHEET_NAME = "Knock-Out"
CRITERIA_COL = 12  # Column L
LABEL_COL = 4      # Column D
HEADER_SCAN_ROWS = 12
YELLOW_FILL = PatternFill(fill_type="solid", start_color="FFFF00", end_color="FFFF00")


def _norm(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip().lower()


def _num(value: object) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value)
    m = re.search(r"-?\d+(?:\.\d+)?", text.replace(",", ""))
    return float(m.group(0)) if m else None


def _is_non_positive_note(value: object) -> bool:
    text = _norm(value)
    return text in {"", "no", "none", "n/a", "na", "nil", "0", "no record"}


def _extract_mia_count(text: str, key: str) -> int:
    m = re.search(rf"{re.escape(key)}\s*[:=]\s*(\d+)", text, re.IGNORECASE)
    return int(m.group(1)) if m else 0


def _extract_mia_count_at_or_above_in_segment(text: str, segment: str, min_level: int) -> int:
    """Sum MIA counts at or above a level within a segment (e.g., MIA2+ in past 6 months)."""
    seg_pattern = re.compile(
        rf"{re.escape(segment)}(.*?)(?:and\s*/?or|$)",
        re.IGNORECASE | re.DOTALL,
    )
    m = seg_pattern.search(text)
    scoped = m.group(1) if m else text

    total = 0
    for level in range(min_level, 5):
        key = f"mia{level}"
        total += _extract_mia_count(scoped, key)

    # Include any "MIA4+" style notation when min_level <= 4
    if min_level <= 4:
        plus_match = re.search(r"mia\s*4\s*\+\s*[:=]\s*(\d+)", scoped, re.IGNORECASE)
        if plus_match:
            total = total - _extract_mia_count(scoped, "mia4") + int(plus_match.group(1))

    return total


def _matches_criteria(criteria: str, value: object, label: str) -> bool:
    c = _norm(criteria)
    v = _norm(value)
    n = _num(value)

    if not c:
        return False

    if "no score" in c and "e or worse" in c:
        if not v:
            return True
        if v in {"e", "f"}:
            return True
        if n is not None and n <= 500:
            return True
        return False

    if c == "no":
        if "business has been in operations" in _norm(label):
            if n is not None:
                return n < 3
        return v.startswith("no") or v in {"0", "false"}

    if "other than \"existing\"" in c:
        return v != "existing"

    if c == "yes":
        return v.startswith("yes")

    if "mia2" in c and "mia1" in c:
        # Rule interpretation:
        # - "past 6 months > 2 times MIA2" means MIA2 and above (MIA2+)
        # - "current 1 month > 4 times MIA1" means MIA1 and above (MIA1+)
        past_6_mia2_plus = _extract_mia_count_at_or_above_in_segment(v, "past 6 months", 2)
        current_1_mia1_plus = _extract_mia_count_at_or_above_in_segment(v, "current 1 month", 1)
        return past_6_mia2_plus > 2 or current_1_mia1_plus > 4

    if c.startswith("≥"):
        threshold = _num(c)
        return threshold is not None and n is not None and n >= threshold

    if c.startswith(">"):
        threshold = _num(c)
        return threshold is not None and n is not None and n > threshold

    if c.startswith("<"):
        threshold = _num(c)
        return threshold is not None and n is not None and n < threshold

    if "subject as \"defendant\"" in c and "ongoing" in c:
        return "defendant" in v and "ongoing" in v

    if "any positive notation" in c or c == "positive":
        return not _is_non_positive_note(value)

    if ">1 & outstanding >rm10k" in c:
        amount = _num(value)
        has_count = bool(re.search(r"\b[2-9]\b", v))
        return has_count and amount is not None and amount > 10000

    # Unhandled criteria: leave unhighlighted.
    return False


def _subject_columns(ws: Worksheet, start_col: int = 13) -> list[int]:
    cols: list[int] = []
    for c in range(start_col, ws.max_column + 1):
        text = " ".join(_norm(ws.cell(r, c).value) for r in range(1, HEADER_SCAN_ROWS + 1))
        if any(k in text for k in ("issuer", "director", "guarantor", "key person")):
            cols.append(c)
    return cols


def highlight_matches(file_path: str, output_path: Optional[str] = None) -> str:
    wb = openpyxl.load_workbook(file_path)
    ws = wb[SHEET_NAME]

    subjects = _subject_columns(ws)
    highlighted = 0

    for row in range(11, ws.max_row + 1):
        label = ws.cell(row, LABEL_COL).value
        criteria = ws.cell(row, CRITERIA_COL).value
        if not criteria:
            continue
        for col in subjects:
            cell = ws.cell(row, col)
            if _matches_criteria(str(criteria), cell.value, str(label or "")):
                cell.fill = YELLOW_FILL
                highlighted += 1

    if not output_path:
        src = Path(file_path)
        output_path = str(src.with_name(f"{src.stem}_COLUMN_L_CHECKED{src.suffix}"))

    wb.save(output_path)
    print(f"✅ Column L validation complete. Highlighted {highlighted} cell(s).")
    print(f"📄 Saved file: {output_path}")
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Check Knock-Out values against Column L criteria and highlight matching cells.")
    parser.add_argument("--excel", required=True, help="Path to filled Knock-Out Excel file")
    parser.add_argument("--output", help="Optional output path")
    args = parser.parse_args()

    highlight_matches(args.excel, args.output)


if __name__ == "__main__":
    main()
